"""PIC 2026 - Plan Industriel et Commercial API"""
from fastapi import APIRouter, Query, HTTPException
from fastapi.responses import StreamingResponse
from typing import Optional
import time
import io

from ..database_unified import execute_app as execute_query

router = APIRouter(prefix="/api/pic", tags=["PIC 2026"])

# Table source
TABLE_CA = "[GROUPE_ALBOUGHAZE].[dbo].[CA_Groupe]"


def safe_float(value, default=0):
    """Convertit une valeur en float de maniere securisee"""
    try:
        return float(value) if value is not None else default
    except (ValueError, TypeError):
        return default


def safe_int(value, default=0):
    """Convertit une valeur en int de maniere securisee"""
    try:
        return int(value) if value is not None else default
    except (ValueError, TypeError):
        return default


@router.get("/2026")
async def get_pic_2026(
    taux_croissance: float = Query(0.05, ge=0, le=0.5, description="Taux de croissance (ex: 0.05 pour 5%)")
):
    """
    Recupere toutes les donnees du PIC 2026 basees sur la procedure stockee sp_PIC_2026.
    Retourne les donnees structurees pour le frontend.
    """
    try:
        start_time = time.time()

        # 1. Recuperer le nombre d'annees
        nb_annees_query = f"""
        SELECT COUNT(DISTINCT [Année]) AS NbAnnees
        FROM {TABLE_CA}
        WHERE [Exclu] = 0 OR [Exclu] IS NULL
        """
        nb_annees_result = execute_query(nb_annees_query)
        nb_annees = safe_int(nb_annees_result[0].get('NbAnnees', 1) if nb_annees_result else 1, 1)
        if nb_annees == 0:
            nb_annees = 1  # Protection division par zero

        # 2. Synthese globale (Section 9.3 de la procedure)
        synthese_query = f"""
        SELECT
            COUNT(DISTINCT [Société]) AS Nb_Societes,
            COUNT(DISTINCT [Région]) AS Nb_Regions,
            COUNT(DISTINCT [Représentant]) AS Nb_Commerciaux,
            COUNT(DISTINCT [Code article]) AS Nb_Articles,
            SUM([Montant HT Net]) AS Total_CA,
            SUM([Montant HT Net]) - SUM([Coût]) AS Total_Marge
        FROM {TABLE_CA}
        WHERE [Exclu] = 0 OR [Exclu] IS NULL
        """
        synthese_result = execute_query(synthese_query)
        synthese_data = synthese_result[0] if synthese_result else {}

        total_ca = safe_float(synthese_data.get('Total_CA', 0))
        total_marge = safe_float(synthese_data.get('Total_Marge', 0))

        synthese = {
            "nb_societes": safe_int(synthese_data.get('Nb_Societes', 0)),
            "nb_regions": safe_int(synthese_data.get('Nb_Regions', 0)),
            "nb_commerciaux": safe_int(synthese_data.get('Nb_Commerciaux', 0)),
            "nb_articles": safe_int(synthese_data.get('Nb_Articles', 0)),
            "nb_annees_historique": nb_annees,
            "ca_annuel_moyen": round(total_ca / nb_annees, 2),
            "objectif_ca_2026": round((total_ca / nb_annees) * (1 + taux_croissance), 2),
            "marge_annuelle_moyenne": round(total_marge / nb_annees, 2),
            "objectif_marge_2026": round((total_marge / nb_annees) * (1 + taux_croissance), 2),
            "taux_croissance": taux_croissance * 100
        }

        # 3. Indices de saisonnalite (Section 1.2)
        saisonnalite_query = f"""
        WITH MoyenneMensuelle AS (
            SELECT
                MONTH([Date BL]) AS Mois,
                [Année],
                SUM([Montant HT Net]) AS CAMois
            FROM {TABLE_CA}
            WHERE [Exclu] = 0 OR [Exclu] IS NULL
            GROUP BY [Année], MONTH([Date BL])
        ),
        SaisonnaliteCalc AS (
            SELECT
                Mois,
                AVG(CAMois) AS CAMoyenMois
            FROM MoyenneMensuelle
            GROUP BY Mois
        )
        SELECT
            Mois,
            DATENAME(MONTH, DATEFROMPARTS(2026, Mois, 1)) AS NomMois,
            ROUND(CAMoyenMois, 2) AS CAMoyenHistorique,
            ROUND(CAMoyenMois / NULLIF((SELECT AVG(CAMoyenMois) FROM SaisonnaliteCalc), 0) * 100, 2) AS IndiceSaisonnalite
        FROM SaisonnaliteCalc
        ORDER BY Mois
        """
        saisonnalite_result = execute_query(saisonnalite_query)
        saisonnalite = [
            {
                "mois": row.get('NomMois', '')[:3] if row.get('NomMois') else f"M{row.get('Mois', 0)}",
                "mois_num": safe_int(row.get('Mois', 0)),
                "ca_moyen": safe_float(row.get('CAMoyenHistorique', 0)),
                "indice": safe_float(row.get('IndiceSaisonnalite', 100))
            }
            for row in saisonnalite_result
        ]

        # 4. Objectifs par Region (Section 4.1 et 7)
        regions_query = f"""
        SELECT
            [Région],
            COUNT(DISTINCT [Code article]) AS NbArticles,
            COUNT(DISTINCT [Représentant]) AS NbCommerciaux,
            SUM([Quantité]) AS TotalQuantite,
            SUM([Montant HT Net]) AS TotalCA,
            SUM([Coût]) AS TotalCout,
            SUM([Montant HT Net]) - SUM([Coût]) AS MargeBrute,
            ROUND((SUM([Montant HT Net]) - SUM([Coût])) / NULLIF(SUM([Montant HT Net]), 0) * 100, 2) AS TauxMarge,
            ROUND(SUM([Montant HT Net]) * 100.0 / SUM(SUM([Montant HT Net])) OVER(), 2) AS PartCA_Percent
        FROM {TABLE_CA}
        WHERE [Exclu] = 0 OR [Exclu] IS NULL
        GROUP BY [Région]
        ORDER BY TotalCA DESC
        """
        regions_result = execute_query(regions_query)
        regions = [
            {
                "region": row.get('Région', 'N/A') or 'N/A',
                "nb_articles": safe_int(row.get('NbArticles', 0)),
                "nb_commerciaux": safe_int(row.get('NbCommerciaux', 0)),
                "quantite": safe_int(row.get('TotalQuantite', 0)),
                "ca": safe_float(row.get('TotalCA', 0)),
                "ca_reference": round(safe_float(row.get('TotalCA', 0)) / nb_annees, 2),
                "objectif_ca_2026": round((safe_float(row.get('TotalCA', 0)) / nb_annees) * (1 + taux_croissance), 2),
                "objectif_qte_2026": round((safe_float(row.get('TotalQuantite', 0)) / nb_annees) * (1 + taux_croissance), 0),
                "marge": safe_float(row.get('MargeBrute', 0)),
                "objectif_marge_2026": round((safe_float(row.get('MargeBrute', 0)) / nb_annees) * (1 + taux_croissance), 2),
                "taux_marge": safe_float(row.get('TauxMarge', 0)),
                "part_ca_percent": safe_float(row.get('PartCA_Percent', 0))
            }
            for row in regions_result
        ]

        # 5. Objectifs par Commercial (Section 4.2 et 8)
        commerciaux_query = f"""
        SELECT
            [Représentant],
            [Région],
            COUNT(DISTINCT [Code article]) AS NbArticles,
            SUM([Quantité]) AS TotalQuantite,
            SUM([Montant HT Net]) AS TotalCA,
            SUM([Montant HT Net]) - SUM([Coût]) AS MargeBrute,
            ROUND((SUM([Montant HT Net]) - SUM([Coût])) / NULLIF(SUM([Montant HT Net]), 0) * 100, 2) AS TauxMarge,
            RANK() OVER (ORDER BY SUM([Montant HT Net]) - SUM([Coût]) DESC) AS RangMarge,
            RANK() OVER (ORDER BY SUM([Montant HT Net]) DESC) AS RangCA
        FROM {TABLE_CA}
        WHERE [Exclu] = 0 OR [Exclu] IS NULL
        GROUP BY [Représentant], [Région]
        ORDER BY TotalCA DESC
        """
        commerciaux_result = execute_query(commerciaux_query)
        commerciaux = [
            {
                "representant": row.get('Représentant', 'N/A') or 'N/A',
                "region": row.get('Région', 'N/A') or 'N/A',
                "nb_articles": safe_int(row.get('NbArticles', 0)),
                "quantite": safe_int(row.get('TotalQuantite', 0)),
                "ca": safe_float(row.get('TotalCA', 0)),
                "ca_reference": round(safe_float(row.get('TotalCA', 0)) / nb_annees, 2),
                "objectif_ca_2026": round((safe_float(row.get('TotalCA', 0)) / nb_annees) * (1 + taux_croissance), 2),
                "objectif_qte_2026": round((safe_float(row.get('TotalQuantite', 0)) / nb_annees) * (1 + taux_croissance), 0),
                "marge": safe_float(row.get('MargeBrute', 0)),
                "objectif_marge_2026": round((safe_float(row.get('MargeBrute', 0)) / nb_annees) * (1 + taux_croissance), 2),
                "taux_marge": safe_float(row.get('TauxMarge', 0)),
                "rang_ca": safe_int(row.get('RangCA', 0)),
                "rang_marge": safe_int(row.get('RangMarge', 0))
            }
            for row in commerciaux_result
        ]

        # 6. Classification ABC des articles (Section 5.1)
        articles_query = f"""
        WITH ArticleCA AS (
            SELECT
                [Code article],
                MAX([Désignation]) AS Designation,
                MAX([Catalogue 1]) AS Catalogue1,
                MAX([Catalogue 2]) AS Catalogue2,
                MAX([Catalogue 3]) AS Catalogue3,
                SUM([Montant HT Net]) AS CA_Article,
                SUM([Quantité]) AS Qte_Article,
                SUM([Coût]) AS Cout_Article,
                SUM([Montant HT Net]) - SUM([Coût]) AS Marge_Article,
                COUNT(DISTINCT [Région]) AS NbRegions
            FROM {TABLE_CA}
            WHERE [Exclu] = 0 OR [Exclu] IS NULL
            GROUP BY [Code article]
        ),
        ArticleRank AS (
            SELECT
                *,
                SUM(CA_Article) OVER (ORDER BY CA_Article DESC ROWS UNBOUNDED PRECEDING) AS CA_Cumule,
                SUM(CA_Article) OVER () AS CA_Total
            FROM ArticleCA
        )
        SELECT TOP 100
            [Code article],
            Designation,
            Catalogue1,
            Catalogue2,
            Catalogue3,
            CA_Article,
            Qte_Article,
            Cout_Article,
            Marge_Article,
            NbRegions,
            ROUND(CA_Article / NULLIF(CA_Total, 0) * 100, 2) AS PartCA_Percent,
            ROUND(CA_Cumule / NULLIF(CA_Total, 0) * 100, 2) AS CA_Cumule_Percent,
            CASE
                WHEN CA_Cumule / NULLIF(CA_Total, 0) <= 0.80 THEN 'A'
                WHEN CA_Cumule / NULLIF(CA_Total, 0) <= 0.95 THEN 'B'
                ELSE 'C'
            END AS Classe_ABC,
            ROUND((Marge_Article) / NULLIF(CA_Article, 0) * 100, 2) AS TauxMarge
        FROM ArticleRank
        ORDER BY CA_Article DESC
        """
        articles_result = execute_query(articles_query)
        articles = [
            {
                "code_article": row.get('Code article', ''),
                "designation": row.get('Designation', ''),
                "catalogue_1": row.get('Catalogue1', ''),
                "catalogue_2": row.get('Catalogue2', ''),
                "catalogue_3": row.get('Catalogue3', ''),
                "ca": safe_float(row.get('CA_Article', 0)),
                "qte_vendue": safe_int(row.get('Qte_Article', 0)),
                "cout": safe_float(row.get('Cout_Article', 0)),
                "marge": safe_float(row.get('Marge_Article', 0)),
                "taux_marge": safe_float(row.get('TauxMarge', 0)),
                "nb_regions": safe_int(row.get('NbRegions', 0)),
                "part_ca_percent": safe_float(row.get('PartCA_Percent', 0)),
                "ca_cumule_percent": safe_float(row.get('CA_Cumule_Percent', 0)),
                "classe": row.get('Classe_ABC', 'C'),
                "qte_prevue_2026": round((safe_float(row.get('Qte_Article', 0)) / nb_annees) * (1 + taux_croissance), 0)
            }
            for row in articles_result
        ]

        # 7. Objectifs mensuels 2026 (Section 4.3)
        objectifs_mensuels_query = f"""
        WITH Saisonnalite AS (
            SELECT
                MONTH([Date BL]) AS Mois,
                SUM([Montant HT Net]) AS CA_Mois,
                SUM([Montant HT Net]) - SUM([Coût]) AS Marge_Mois
            FROM {TABLE_CA}
            WHERE [Exclu] = 0 OR [Exclu] IS NULL
            GROUP BY MONTH([Date BL])
        ),
        Totaux AS (
            SELECT
                SUM(CA_Mois) AS CA_Total,
                SUM(Marge_Mois) AS Marge_Total
            FROM Saisonnalite
        )
        SELECT
            s.Mois,
            DATENAME(MONTH, DATEFROMPARTS(2026, s.Mois, 1)) AS NomMois,
            ROUND(s.CA_Mois / NULLIF(t.CA_Total, 0) * 100, 2) AS PoidsMonth_Percent,
            ROUND((t.CA_Total / {nb_annees}) * (s.CA_Mois / NULLIF(t.CA_Total, 0)) * (1 + {taux_croissance}), 2) AS Objectif_CA_Mensuel,
            ROUND((t.Marge_Total / {nb_annees}) * (s.Marge_Mois / NULLIF(t.Marge_Total, 0)) * (1 + {taux_croissance}), 2) AS Objectif_Marge_Mensuel
        FROM Saisonnalite s
        CROSS JOIN Totaux t
        ORDER BY s.Mois
        """
        objectifs_mensuels_result = execute_query(objectifs_mensuels_query)
        objectifs_mensuels = [
            {
                "mois": row.get('NomMois', ''),
                "mois_num": safe_int(row.get('Mois', 0)),
                "poids_percent": safe_float(row.get('PoidsMonth_Percent', 0)),
                "objectif_ca": safe_float(row.get('Objectif_CA_Mensuel', 0)),
                "objectif_marge": safe_float(row.get('Objectif_Marge_Mensuel', 0))
            }
            for row in objectifs_mensuels_result
        ]

        # 8. Analyse par Catalogue (Section 5.2)
        catalogues_query = f"""
        SELECT
            [Catalogue 1] AS Catalogue,
            COUNT(DISTINCT [Code article]) AS NbArticles,
            SUM([Quantité]) AS TotalQuantite,
            SUM([Montant HT Net]) AS TotalCA,
            SUM([Montant HT Net]) - SUM([Coût]) AS TotalMarge,
            ROUND((SUM([Montant HT Net]) - SUM([Coût])) / NULLIF(SUM([Montant HT Net]), 0) * 100, 2) AS TauxMarge,
            ROUND(SUM([Montant HT Net]) * 100.0 / SUM(SUM([Montant HT Net])) OVER(), 2) AS PartCA_Percent
        FROM {TABLE_CA}
        WHERE [Exclu] = 0 OR [Exclu] IS NULL
        GROUP BY [Catalogue 1]
        ORDER BY TotalCA DESC
        """
        catalogues_result = execute_query(catalogues_query)
        catalogues = [
            {
                "catalogue": row.get('Catalogue', 'N/A') or 'N/A',
                "nb_articles": safe_int(row.get('NbArticles', 0)),
                "qte_totale": safe_int(row.get('TotalQuantite', 0)),
                "ca": safe_float(row.get('TotalCA', 0)),
                "marge": safe_float(row.get('TotalMarge', 0)),
                "taux_marge": safe_float(row.get('TauxMarge', 0)),
                "part_ca_percent": safe_float(row.get('PartCA_Percent', 0))
            }
            for row in catalogues_result
        ]

        # 9. Recommandations Charge/Capacite (Section 9.1)
        charge_query = f"""
        DECLARE @MaxAnnee INT;
        SELECT @MaxAnnee = MAX([Année]) FROM {TABLE_CA} WHERE [Exclu] = 0 OR [Exclu] IS NULL;

        DECLARE @Qte_Annuelle DECIMAL(18,2);
        SELECT @Qte_Annuelle = SUM([Quantité])
        FROM {TABLE_CA}
        WHERE [Année] = @MaxAnnee AND ([Exclu] = 0 OR [Exclu] IS NULL);

        SELECT
            MONTH([Date BL]) AS Mois,
            DATENAME(MONTH, DATEFROMPARTS(2026, MONTH([Date BL]), 1)) AS NomMois,
            SUM([Quantité]) AS Qte_Totale,
            ROUND(SUM([Quantité]) / NULLIF(@Qte_Annuelle, 0) * 100, 2) AS Poids_Percent,
            ROUND(SUM([Quantité]) * (1 + {taux_croissance}), 0) AS Charge_Prevue_2026,
            CASE
                WHEN SUM([Quantité]) / NULLIF(@Qte_Annuelle, 0) > 0.10 THEN 'HAUTE'
                WHEN SUM([Quantité]) / NULLIF(@Qte_Annuelle, 0) < 0.06 THEN 'BASSE'
                ELSE 'NORMALE'
            END AS Periode
        FROM {TABLE_CA}
        WHERE [Année] = @MaxAnnee AND ([Exclu] = 0 OR [Exclu] IS NULL)
        GROUP BY MONTH([Date BL])
        ORDER BY Mois
        """
        charge_result = execute_query(charge_query)
        charge_capacite = [
            {
                "mois": row.get('NomMois', ''),
                "mois_num": safe_int(row.get('Mois', 0)),
                "qte_totale": safe_int(row.get('Qte_Totale', 0)),
                "poids_percent": safe_float(row.get('Poids_Percent', 0)),
                "charge_prevue_2026": safe_int(row.get('Charge_Prevue_2026', 0)),
                "periode": row.get('Periode', 'NORMALE')
            }
            for row in charge_result
        ]

        query_time = time.time() - start_time

        return {
            "success": True,
            "synthese": synthese,
            "saisonnalite": saisonnalite,
            "regions": regions,
            "commerciaux": commerciaux,
            "articles": articles,
            "objectifsMensuels": objectifs_mensuels,
            "catalogues": catalogues,
            "chargeCapacite": charge_capacite,
            "taux_croissance": taux_croissance,
            "query_time_ms": round(query_time * 1000, 2)
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/2026/regions")
async def get_pic_regions(
    taux_croissance: float = Query(0.05, ge=0, le=0.5)
):
    """Recupere les donnees detaillees par region"""
    try:
        # Nombre d'annees
        nb_annees_query = f"""
        SELECT COUNT(DISTINCT [Année]) AS NbAnnees FROM {TABLE_CA} WHERE [Exclu] = 0 OR [Exclu] IS NULL
        """
        nb_annees_result = execute_query(nb_annees_query)
        nb_annees = safe_int(nb_annees_result[0].get('NbAnnees', 1) if nb_annees_result else 1, 1)
        if nb_annees == 0:
            nb_annees = 1

        # Performance par Region avec statut
        query = f"""
        WITH StatsRegion AS (
            SELECT
                [Région],
                COUNT(DISTINCT [Code article]) AS NbArticles,
                COUNT(DISTINCT [Représentant]) AS NbCommerciaux,
                SUM([Quantité]) AS TotalQuantite,
                SUM([Montant HT Net]) AS TotalCA,
                SUM([Montant HT Net]) - SUM([Coût]) AS MargeBrute,
                ROUND((SUM([Montant HT Net]) - SUM([Coût])) / NULLIF(SUM([Montant HT Net]), 0) * 100, 2) AS TauxMarge,
                ROUND(SUM([Montant HT Net]) * 100.0 / SUM(SUM([Montant HT Net])) OVER(), 2) AS PartCA_Percent
            FROM {TABLE_CA}
            WHERE [Exclu] = 0 OR [Exclu] IS NULL
            GROUP BY [Région]
        )
        SELECT
            *,
            (SELECT AVG(TotalCA) FROM StatsRegion) AS CA_Moyen_Regions,
            CASE
                WHEN TotalCA < (SELECT AVG(TotalCA) FROM StatsRegion) THEN 'SOUS-PERFORMANCE'
                WHEN TotalCA >= (SELECT AVG(TotalCA) FROM StatsRegion) * 1.2 THEN 'SURPERFORMANCE'
                ELSE 'DANS LA MOYENNE'
            END AS Statut
        FROM StatsRegion
        ORDER BY TotalCA DESC
        """
        result = execute_query(query)

        regions = [
            {
                "region": row.get('Région', 'N/A') or 'N/A',
                "nb_articles": safe_int(row.get('NbArticles', 0)),
                "nb_commerciaux": safe_int(row.get('NbCommerciaux', 0)),
                "quantite": safe_int(row.get('TotalQuantite', 0)),
                "ca": safe_float(row.get('TotalCA', 0)),
                "ca_reference": round(safe_float(row.get('TotalCA', 0)) / nb_annees, 2),
                "objectif_ca_2026": round((safe_float(row.get('TotalCA', 0)) / nb_annees) * (1 + taux_croissance), 2),
                "objectif_qte_2026": round((safe_float(row.get('TotalQuantite', 0)) / nb_annees) * (1 + taux_croissance), 0),
                "marge": safe_float(row.get('MargeBrute', 0)),
                "objectif_marge_2026": round((safe_float(row.get('MargeBrute', 0)) / nb_annees) * (1 + taux_croissance), 2),
                "taux_marge": safe_float(row.get('TauxMarge', 0)),
                "part_ca_percent": safe_float(row.get('PartCA_Percent', 0)),
                "ca_moyen_regions": safe_float(row.get('CA_Moyen_Regions', 0)),
                "statut": row.get('Statut', 'DANS LA MOYENNE')
            }
            for row in result
        ]

        return {"success": True, "data": regions}

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/2026/commerciaux")
async def get_pic_commerciaux(
    taux_croissance: float = Query(0.05, ge=0, le=0.5),
    region: Optional[str] = Query(None, description="Filtre par region")
):
    """Recupere les donnees detaillees par commercial"""
    try:
        # Nombre d'annees
        nb_annees_query = f"""
        SELECT COUNT(DISTINCT [Année]) AS NbAnnees FROM {TABLE_CA} WHERE [Exclu] = 0 OR [Exclu] IS NULL
        """
        nb_annees_result = execute_query(nb_annees_query)
        nb_annees = safe_int(nb_annees_result[0].get('NbAnnees', 1) if nb_annees_result else 1, 1)
        if nb_annees == 0:
            nb_annees = 1

        # Filtres
        where_clauses = ["([Exclu] = 0 OR [Exclu] IS NULL)"]
        params = []
        if region:
            where_clauses.append("[Région] = ?")
            params.append(region)

        where_sql = " AND ".join(where_clauses)

        query = f"""
        SELECT
            [Représentant],
            [Région],
            COUNT(DISTINCT [Code article]) AS NbArticles,
            SUM([Quantité]) AS TotalQuantite,
            SUM([Montant HT Net]) AS TotalCA,
            SUM([Montant HT Net]) - SUM([Coût]) AS MargeBrute,
            ROUND((SUM([Montant HT Net]) - SUM([Coût])) / NULLIF(SUM([Montant HT Net]), 0) * 100, 2) AS TauxMarge,
            RANK() OVER (ORDER BY SUM([Montant HT Net]) - SUM([Coût]) DESC) AS RangMarge,
            RANK() OVER (ORDER BY SUM([Montant HT Net]) DESC) AS RangCA
        FROM {TABLE_CA}
        WHERE {where_sql}
        GROUP BY [Représentant], [Région]
        ORDER BY TotalCA DESC
        """
        result = execute_query(query, tuple(params) if params else None)

        commerciaux = [
            {
                "representant": row.get('Représentant', 'N/A') or 'N/A',
                "region": row.get('Région', 'N/A') or 'N/A',
                "nb_articles": safe_int(row.get('NbArticles', 0)),
                "quantite": safe_int(row.get('TotalQuantite', 0)),
                "ca": safe_float(row.get('TotalCA', 0)),
                "ca_reference": round(safe_float(row.get('TotalCA', 0)) / nb_annees, 2),
                "objectif_ca_2026": round((safe_float(row.get('TotalCA', 0)) / nb_annees) * (1 + taux_croissance), 2),
                "objectif_qte_2026": round((safe_float(row.get('TotalQuantite', 0)) / nb_annees) * (1 + taux_croissance), 0),
                "marge": safe_float(row.get('MargeBrute', 0)),
                "objectif_marge_2026": round((safe_float(row.get('MargeBrute', 0)) / nb_annees) * (1 + taux_croissance), 2),
                "taux_marge": safe_float(row.get('TauxMarge', 0)),
                "rang_ca": safe_int(row.get('RangCA', 0)),
                "rang_marge": safe_int(row.get('RangMarge', 0))
            }
            for row in result
        ]

        return {"success": True, "data": commerciaux}

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/2026/articles")
async def get_pic_articles(
    taux_croissance: float = Query(0.05, ge=0, le=0.5),
    catalogue: Optional[str] = Query(None, description="Filtre par catalogue 1"),
    classe_abc: Optional[str] = Query(None, description="Filtre par classe ABC (A, B, C)"),
    limit: int = Query(100, ge=10, le=500)
):
    """Recupere les donnees detaillees par article avec classification ABC"""
    try:
        # Nombre d'annees
        nb_annees_query = f"""
        SELECT COUNT(DISTINCT [Année]) AS NbAnnees FROM {TABLE_CA} WHERE [Exclu] = 0 OR [Exclu] IS NULL
        """
        nb_annees_result = execute_query(nb_annees_query)
        nb_annees = safe_int(nb_annees_result[0].get('NbAnnees', 1) if nb_annees_result else 1, 1)
        if nb_annees == 0:
            nb_annees = 1

        # Construction des filtres pour la CTE
        where_base = "[Exclu] = 0 OR [Exclu] IS NULL"
        if catalogue:
            where_base += f" AND [Catalogue 1] = '{catalogue}'"

        query = f"""
        WITH ArticleCA AS (
            SELECT
                [Code article],
                MAX([Désignation]) AS Designation,
                MAX([Catalogue 1]) AS Catalogue1,
                MAX([Catalogue 2]) AS Catalogue2,
                MAX([Catalogue 3]) AS Catalogue3,
                SUM([Montant HT Net]) AS CA_Article,
                SUM([Quantité]) AS Qte_Article,
                SUM([Coût]) AS Cout_Article,
                SUM([Montant HT Net]) - SUM([Coût]) AS Marge_Article,
                COUNT(DISTINCT [Région]) AS NbRegions
            FROM {TABLE_CA}
            WHERE {where_base}
            GROUP BY [Code article]
        ),
        ArticleRank AS (
            SELECT
                *,
                SUM(CA_Article) OVER (ORDER BY CA_Article DESC ROWS UNBOUNDED PRECEDING) AS CA_Cumule,
                SUM(CA_Article) OVER () AS CA_Total
            FROM ArticleCA
        ),
        ArticleClassified AS (
            SELECT
                *,
                ROUND(CA_Article / NULLIF(CA_Total, 0) * 100, 2) AS PartCA_Percent,
                ROUND(CA_Cumule / NULLIF(CA_Total, 0) * 100, 2) AS CA_Cumule_Percent,
                CASE
                    WHEN CA_Cumule / NULLIF(CA_Total, 0) <= 0.80 THEN 'A'
                    WHEN CA_Cumule / NULLIF(CA_Total, 0) <= 0.95 THEN 'B'
                    ELSE 'C'
                END AS Classe_ABC,
                ROUND((Marge_Article) / NULLIF(CA_Article, 0) * 100, 2) AS TauxMarge
            FROM ArticleRank
        )
        SELECT TOP {limit} * FROM ArticleClassified
        {"WHERE Classe_ABC = '" + classe_abc + "'" if classe_abc else ""}
        ORDER BY CA_Article DESC
        """
        result = execute_query(query)

        articles = [
            {
                "code_article": row.get('Code article', ''),
                "designation": row.get('Designation', ''),
                "catalogue_1": row.get('Catalogue1', ''),
                "catalogue_2": row.get('Catalogue2', ''),
                "catalogue_3": row.get('Catalogue3', ''),
                "ca": safe_float(row.get('CA_Article', 0)),
                "qte_vendue": safe_int(row.get('Qte_Article', 0)),
                "cout": safe_float(row.get('Cout_Article', 0)),
                "marge": safe_float(row.get('Marge_Article', 0)),
                "taux_marge": safe_float(row.get('TauxMarge', 0)),
                "nb_regions": safe_int(row.get('NbRegions', 0)),
                "part_ca_percent": safe_float(row.get('PartCA_Percent', 0)),
                "ca_cumule_percent": safe_float(row.get('CA_Cumule_Percent', 0)),
                "classe": row.get('Classe_ABC', 'C'),
                "qte_prevue_2026": round((safe_float(row.get('Qte_Article', 0)) / nb_annees) * (1 + taux_croissance), 0),
                "objectif_ca_2026": round((safe_float(row.get('CA_Article', 0)) / nb_annees) * (1 + taux_croissance), 2)
            }
            for row in result
        ]

        return {"success": True, "data": articles}

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/2026/articles-faible-marge")
async def get_articles_faible_marge(
    seuil_marge: float = Query(10, description="Seuil de marge en % (defaut 10%)"),
    seuil_ca: float = Query(1000, description="Seuil minimum de CA")
):
    """Recupere les articles a faible marge (< seuil) avec un CA significatif"""
    try:
        query = f"""
        SELECT
            [Code article],
            MAX([Désignation]) AS Designation,
            MAX([Catalogue 1]) AS Catalogue1,
            SUM([Montant HT Net]) AS CA,
            SUM([Coût]) AS Cout,
            SUM([Montant HT Net]) - SUM([Coût]) AS Marge,
            ROUND((SUM([Montant HT Net]) - SUM([Coût])) / NULLIF(SUM([Montant HT Net]), 0) * 100, 2) AS TauxMarge
        FROM {TABLE_CA}
        WHERE [Exclu] = 0 OR [Exclu] IS NULL
        GROUP BY [Code article]
        HAVING (SUM([Montant HT Net]) - SUM([Coût])) / NULLIF(SUM([Montant HT Net]), 0) * 100 < {seuil_marge}
           AND SUM([Montant HT Net]) > {seuil_ca}
        ORDER BY CA DESC
        """
        result = execute_query(query)

        articles = [
            {
                "code_article": row.get('Code article', ''),
                "designation": row.get('Designation', ''),
                "catalogue_1": row.get('Catalogue1', ''),
                "ca": safe_float(row.get('CA', 0)),
                "cout": safe_float(row.get('Cout', 0)),
                "marge": safe_float(row.get('Marge', 0)),
                "taux_marge": safe_float(row.get('TauxMarge', 0))
            }
            for row in result
        ]

        return {"success": True, "data": articles, "seuil_marge": seuil_marge, "seuil_ca": seuil_ca}

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/2026/articles-marge-negative")
async def get_articles_marge_negative():
    """Recupere les articles a marge negative (en perte)"""
    try:
        query = f"""
        SELECT
            [Code article],
            MAX([Désignation]) AS Designation,
            MAX([Catalogue 1]) AS Catalogue1,
            SUM([Quantité]) AS Qte,
            SUM([Montant HT Net]) AS CA,
            SUM([Coût]) AS Cout,
            SUM([Montant HT Net]) - SUM([Coût]) AS Marge_Negative
        FROM {TABLE_CA}
        WHERE [Exclu] = 0 OR [Exclu] IS NULL
        GROUP BY [Code article]
        HAVING SUM([Montant HT Net]) - SUM([Coût]) < 0
        ORDER BY Marge_Negative ASC
        """
        result = execute_query(query)

        articles = [
            {
                "code_article": row.get('Code article', ''),
                "designation": row.get('Designation', ''),
                "catalogue_1": row.get('Catalogue1', ''),
                "qte": safe_int(row.get('Qte', 0)),
                "ca": safe_float(row.get('CA', 0)),
                "cout": safe_float(row.get('Cout', 0)),
                "marge": safe_float(row.get('Marge_Negative', 0))
            }
            for row in result
        ]

        return {"success": True, "data": articles}

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/2026/export")
async def export_pic_2026(
    taux_croissance: float = Query(0.05, ge=0, le=0.5)
):
    """Exporte le PIC 2026 en format Excel"""
    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows

        # Recuperer les donnees
        data = await get_pic_2026(taux_croissance)

        # Creer le workbook
        wb = Workbook()

        # Style header
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Feuille Synthese
        ws_synthese = wb.active
        ws_synthese.title = "Synthese"
        synthese = data['synthese']
        ws_synthese.append(["PIC 2026 - SYNTHESE", ""])
        ws_synthese.append(["Taux de croissance", f"{synthese['taux_croissance']}%"])
        ws_synthese.append(["Nombre d'annees historiques", synthese['nb_annees_historique']])
        ws_synthese.append(["Nombre de societes", synthese['nb_societes']])
        ws_synthese.append(["Nombre de regions", synthese['nb_regions']])
        ws_synthese.append(["Nombre de commerciaux", synthese['nb_commerciaux']])
        ws_synthese.append(["Nombre d'articles", synthese['nb_articles']])
        ws_synthese.append([""])
        ws_synthese.append(["CA Annuel Moyen", synthese['ca_annuel_moyen']])
        ws_synthese.append(["Objectif CA 2026", synthese['objectif_ca_2026']])
        ws_synthese.append(["Marge Annuelle Moyenne", synthese['marge_annuelle_moyenne']])
        ws_synthese.append(["Objectif Marge 2026", synthese['objectif_marge_2026']])

        # Feuille Regions
        ws_regions = wb.create_sheet("Regions")
        df_regions = pd.DataFrame(data['regions'])
        for r_idx, row in enumerate(dataframe_to_rows(df_regions, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_regions.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                cell.border = thin_border

        # Feuille Commerciaux
        ws_commerciaux = wb.create_sheet("Commerciaux")
        df_commerciaux = pd.DataFrame(data['commerciaux'])
        for r_idx, row in enumerate(dataframe_to_rows(df_commerciaux, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_commerciaux.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                cell.border = thin_border

        # Feuille Articles
        ws_articles = wb.create_sheet("Articles")
        df_articles = pd.DataFrame(data['articles'])
        for r_idx, row in enumerate(dataframe_to_rows(df_articles, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_articles.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                cell.border = thin_border

        # Feuille Objectifs Mensuels
        ws_objectifs = wb.create_sheet("Objectifs Mensuels")
        df_objectifs = pd.DataFrame(data['objectifsMensuels'])
        for r_idx, row in enumerate(dataframe_to_rows(df_objectifs, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_objectifs.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                cell.border = thin_border

        # Sauvegarder dans un buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=PIC_2026_{int(taux_croissance*100)}pct.xlsx"}
        )

    except ImportError:
        raise HTTPException(status_code=500, detail="Module pandas ou openpyxl non disponible")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
