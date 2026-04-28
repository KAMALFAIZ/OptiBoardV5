"""Routes pour l'Excel Builder - Creation de rapports Excel personnalises"""
import json
import logging
import os
import tempfile
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Header, HTTPException
from fastapi.responses import FileResponse
from pydantic import BaseModel

from ..database_unified import (
    central_cursor as get_db_cursor,
    client_manager,
    current_dwh_code as _ctx_dwh,
    execute_central as execute_query,
    execute_client,
    write_client,
)

logger = logging.getLogger("ExcelBuilder")

# Guard : les DDL de init_excel_builder_tables ne s'executent qu'une seule fois
_eb_tables_initialized = False


# =============================================================================
# HELPERS — Routage client DB / centrale pour APP_ExcelBuilders
# =============================================================================

def _eb_read(query: str, params: tuple = (), dwh_code: str = None) -> list:
    """
    Lit depuis APP_ExcelBuilders avec priorite CENTRALE :
    1. Cherche d'abord en base CENTRALE (OptiBoard_SaaS) — source de verite pour tous les clients
    2. Si pas trouve en centrale → cherche dans la base client (builders specifiques au client)
    """
    # 1. Centrale en priorite
    try:
        central_result = execute_query(query, params, use_cache=False)
        if central_result:
            return central_result
    except Exception as e:
        logger.debug(f"_eb_read central error: {e}")

    # 2. Fallback : base client
    if dwh_code:
        try:
            if client_manager.has_client_db(dwh_code):
                return execute_client(query, params, dwh_code=dwh_code, use_cache=False) or []
        except Exception as e:
            logger.debug(f"_eb_read client fallback ({dwh_code}): {e}")

    return []


def _eb_write(query: str, params: tuple = (), dwh_code: str = None) -> None:
    """
    Ecrit dans APP_ExcelBuilders :
    - Si dwh_code fourni et base client dispo → base client
    - Sinon → base centrale
    """
    if dwh_code:
        try:
            if client_manager.has_client_db(dwh_code):
                write_client(query, params, dwh_code=dwh_code)
                return
        except Exception as e:
            logger.debug(f"_eb_write client fallback ({dwh_code}): {e}")
    with get_db_cursor() as cursor:
        cursor.execute(query, params)


# =============================================================================
# INITIALISATION DE LA TABLE
# =============================================================================

def init_excel_builder_tables() -> bool:
    """Cree la table pour l'Excel Builder — DDL execute une seule fois par demarrage."""
    global _eb_tables_initialized
    if _eb_tables_initialized:
        return True

    create_query = """
    IF NOT EXISTS (SELECT * FROM sys.objects WHERE name='APP_ExcelBuilders')
    CREATE TABLE APP_ExcelBuilders (
        id INT IDENTITY(1,1) PRIMARY KEY,
        name NVARCHAR(200) NOT NULL,
        description NVARCHAR(500) NULL,
        config NVARCHAR(MAX) NOT NULL DEFAULT '{}',
        created_by NVARCHAR(100) NULL,
        created_at DATETIME DEFAULT GETDATE(),
        updated_at DATETIME DEFAULT GETDATE()
    )
    """

    try:
        with get_db_cursor() as cursor:
            cursor.execute(create_query)
        _eb_tables_initialized = True
        logger.info("Table APP_ExcelBuilders initialisee avec succes.")
        return True
    except Exception as e:
        logger.error(f"Erreur init excel builder tables: {e}")
        return False


# =============================================================================
# ROUTER
# =============================================================================

router = APIRouter(prefix="/api/excel-builder", tags=["excel-builder"])


# =============================================================================
# MODELES PYDANTIC
# =============================================================================

class ExcelBuilderCreate(BaseModel):
    name: str
    description: Optional[str] = None
    config: Dict[str, Any] = {}


class ExcelBuilderUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    config: Optional[Dict[str, Any]] = None


# =============================================================================
# ENDPOINT 1 — Liste tous les builders
# =============================================================================

@router.get("/list")
async def list_excel_builders(
    dwh_code: Optional[str] = Header(None, alias="X-DWH-Code"),
):
    """Liste tous les Excel Builders (id, name, description, created_at, updated_at)"""
    try:
        results = _eb_read(
            """SELECT id, name, description, created_by, created_at, updated_at
               FROM APP_ExcelBuilders
               ORDER BY updated_at DESC""",
            dwh_code=dwh_code,
        )
        return {"success": True, "data": results}
    except Exception as e:
        logger.error(f"list_excel_builders: {e}")
        return {"success": False, "error": str(e), "data": []}


# =============================================================================
# ENDPOINT 2 — Recupere un builder avec sa config complete
# =============================================================================

@router.get("/{builder_id}")
async def get_excel_builder(
    builder_id: int,
    dwh_code: Optional[str] = Header(None, alias="X-DWH-Code"),
):
    """Recupere un Excel Builder avec sa configuration complete"""
    try:
        results = _eb_read(
            "SELECT * FROM APP_ExcelBuilders WHERE id = ?",
            (builder_id,),
            dwh_code=dwh_code,
        )
        if not results:
            raise HTTPException(status_code=404, detail="Excel Builder non trouve")

        builder = dict(results[0])
        # Parser le JSON de config
        raw_config = builder.get("config") or "{}"
        builder["config"] = json.loads(raw_config) if isinstance(raw_config, str) else raw_config

        return {"success": True, "data": builder}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"get_excel_builder({builder_id}): {e}")
        raise HTTPException(status_code=500, detail=str(e))


# =============================================================================
# ENDPOINT 3 — Creer un nouveau builder
# =============================================================================

@router.post("/create")
async def create_excel_builder(
    payload: ExcelBuilderCreate,
    dwh_code: Optional[str] = Header(None, alias="X-DWH-Code"),
):
    """Cree un nouvel Excel Builder"""
    try:
        config_json = json.dumps(payload.config, ensure_ascii=False)

        with get_db_cursor() as cursor:
            cursor.execute(
                """INSERT INTO APP_ExcelBuilders (name, description, config, created_by)
                   VALUES (?, ?, ?, ?)""",
                (
                    payload.name,
                    payload.description,
                    config_json,
                    None,  # created_by peut etre alimente via header user si necessaire
                ),
            )
            cursor.execute("SELECT @@IDENTITY AS id")
            row = cursor.fetchone()
            new_id = row[0] if row else None

        return {"success": True, "message": "Excel Builder cree", "id": new_id}
    except Exception as e:
        logger.error(f"create_excel_builder: {e}")
        raise HTTPException(status_code=400, detail=str(e))


# =============================================================================
# ENDPOINT 4 — Mettre a jour un builder
# =============================================================================

@router.put("/{builder_id}")
async def update_excel_builder(
    builder_id: int,
    payload: ExcelBuilderUpdate,
    dwh_code: Optional[str] = Header(None, alias="X-DWH-Code"),
):
    """Met a jour un Excel Builder existant"""
    try:
        updates: List[str] = []
        params: List[Any] = []

        if payload.name is not None:
            updates.append("name = ?")
            params.append(payload.name)
        if payload.description is not None:
            updates.append("description = ?")
            params.append(payload.description)
        if payload.config is not None:
            updates.append("config = ?")
            params.append(json.dumps(payload.config, ensure_ascii=False))

        updates.append("updated_at = GETDATE()")

        if len(updates) == 1:
            # Seul updated_at serait modifie — rien d'utile
            return {"success": False, "message": "Aucune modification fournie"}

        params.append(builder_id)
        query = f"UPDATE APP_ExcelBuilders SET {', '.join(updates)} WHERE id = ?"

        with get_db_cursor() as cursor:
            cursor.execute(query, tuple(params))

        return {"success": True, "message": "Excel Builder mis a jour"}
    except Exception as e:
        logger.error(f"update_excel_builder({builder_id}): {e}")
        raise HTTPException(status_code=400, detail=str(e))


# =============================================================================
# ENDPOINT 5 — Supprimer un builder
# =============================================================================

@router.delete("/{builder_id}")
async def delete_excel_builder(
    builder_id: int,
    dwh_code: Optional[str] = Header(None, alias="X-DWH-Code"),
):
    """Supprime un Excel Builder"""
    try:
        with get_db_cursor() as cursor:
            cursor.execute("DELETE FROM APP_ExcelBuilders WHERE id = ?", (builder_id,))
        return {"success": True, "message": "Excel Builder supprime"}
    except Exception as e:
        logger.error(f"delete_excel_builder({builder_id}): {e}")
        raise HTTPException(status_code=400, detail=str(e))


# =============================================================================
# LOGIQUE D'EXECUTION INTERNE
# =============================================================================

# Noms des mois en francais pour les colonnes
MOIS_LABELS = ["janv", "févr", "mars", "avr", "mai", "juin",
               "juil", "août", "sept", "oct", "nov", "déc"]


def _query_dwh_montant(dwh_code: str, source_table: str, mois: int, annee: int, filters: dict) -> float:
    """
    Requete le DWH pour obtenir SUM(Montant) pour un mois/annee donne.
    Retourne 0.0 en cas d'erreur ou si la table n'existe pas.
    """
    from ..database_unified import DWHConnectionManager

    # Construire les clauses de filtre supplementaires
    filtre_clauses = []
    filtre_params: List[Any] = []

    for col, val in (filters or {}).items():
        # Validation basique du nom de colonne (alphanum + underscore)
        col_safe = "".join(c for c in col if c.isalnum() or c == "_")
        if not col_safe:
            continue
        filtre_clauses.append(f"[{col_safe}] = ?")
        filtre_params.append(val)

    where_extra = (" AND " + " AND ".join(filtre_clauses)) if filtre_clauses else ""

    sql = (
        f"SELECT ISNULL(SUM(Montant), 0) AS val "
        f"FROM [{source_table}] "
        f"WHERE MONTH(DateEcr) = ? AND YEAR(DateEcr) = ?"
        f"{where_extra}"
    )
    params = tuple([mois, annee] + filtre_params)

    try:
        rows = DWHConnectionManager.execute_dwh_query(dwh_code, sql, params=params, use_cache=False)
        if rows:
            val = rows[0].get("val", 0)
            return float(val) if val is not None else 0.0
        return 0.0
    except Exception as e:
        logger.debug(f"_query_dwh_montant table={source_table} mois={mois} annee={annee}: {e}")
        return 0.0


def _execute_builder_logic(config: dict, annee: int, dwh_code: str) -> dict:
    """
    Execute la logique du builder : parcourt les sections/lignes de la config
    et interroge le DWH pour chaque cellule mensuelle.

    Structure config attendue :
    {
        "sections": [
            {
                "label": "Charges",
                "color": "#ff0000",
                "rows": [
                    {
                        "label": "Loyer",
                        "type": "data",          # "data" | "total" | "formula"
                        "source_table": "Ecritures",
                        "filters": {"CodeJournal": "LO"}
                    },
                    {
                        "label": "Total Charges",
                        "type": "total"
                    }
                ]
            }
        ]
    }
    """
    sections_result = []
    colonnes = MOIS_LABELS[:]  # janv..dec

    for section in config.get("sections", []):
        section_label = section.get("label", "")
        section_color = section.get("color", "#2563EB")
        rows_config = section.get("rows", [])

        rows_result = []
        # Stockage des valeurs des lignes "data" pour les totaux
        data_rows_values: List[List[float]] = []

        for row in rows_config:
            row_type = row.get("type", "data")
            row_label = row.get("label", "")

            if row_type == "data":
                source_table = row.get("source_table", "")
                filters = row.get("filters", {})
                valeurs: List[float] = []

                for mois_idx in range(1, 13):
                    if source_table and dwh_code:
                        val = _query_dwh_montant(dwh_code, source_table, mois_idx, annee, filters)
                    else:
                        val = 0.0
                    valeurs.append(val)

                total_ligne = sum(valeurs)
                data_rows_values.append(valeurs)
                rows_result.append({
                    "label": row_label,
                    "type": "data",
                    "values": valeurs,
                    "total": total_ligne,
                })

            elif row_type == "total":
                # Somme de toutes les lignes "data" de la section
                if data_rows_values:
                    valeurs = [sum(row_vals[m] for row_vals in data_rows_values) for m in range(12)]
                else:
                    valeurs = [0.0] * 12
                total_ligne = sum(valeurs)
                rows_result.append({
                    "label": row_label,
                    "type": "total",
                    "values": valeurs,
                    "total": total_ligne,
                })

            elif row_type == "formula":
                formule = row.get("formula", "SUM_SECTION")
                if formule == "SUM_SECTION":
                    if data_rows_values:
                        valeurs = [sum(rv[m] for rv in data_rows_values) for m in range(12)]
                    else:
                        valeurs = [0.0] * 12
                else:
                    valeurs = [0.0] * 12
                total_ligne = sum(valeurs)
                rows_result.append({
                    "label": row_label,
                    "type": "formula",
                    "values": valeurs,
                    "total": total_ligne,
                })

            else:
                # Type inconnu → ligne vide
                rows_result.append({
                    "label": row_label,
                    "type": row_type,
                    "values": [0.0] * 12,
                    "total": 0.0,
                })

        sections_result.append({
            "label": section_label,
            "color": section_color,
            "rows": rows_result,
        })

    return {
        "success": True,
        "columns": colonnes,
        "sections": sections_result,
    }


# =============================================================================
# ENDPOINT 6 — Executer le builder (renvoie les donnees JSON)
# =============================================================================

class ExecuteRequest(BaseModel):
    context: Dict[str, Any] = {}


@router.post("/{builder_id}/execute")
async def execute_excel_builder(
    builder_id: int,
    request: ExecuteRequest = ExecuteRequest(),
    dwh_code: Optional[str] = Header(None, alias="X-DWH-Code"),
):
    """
    Execute un Excel Builder et retourne les donnees structurees par sections/lignes.

    Contexte attendu : { "year": 2025, "dwh_code": "CLIENT01" }
    Le dwh_code dans le contexte est prioritaire sur le header X-DWH-Code.
    """
    try:
        # Recuperer la config du builder
        results = _eb_read(
            "SELECT * FROM APP_ExcelBuilders WHERE id = ?",
            (builder_id,),
            dwh_code=dwh_code,
        )
        if not results:
            raise HTTPException(status_code=404, detail="Excel Builder non trouve")

        builder = results[0]
        raw_config = builder.get("config") or "{}"
        config = json.loads(raw_config) if isinstance(raw_config, str) else raw_config

        # Determiner l'annee et le DWH code effectif
        from datetime import date
        context = request.context or {}
        annee = int(context.get("year", date.today().year))
        effective_dwh = context.get("dwh_code") or dwh_code or _ctx_dwh.get()

        if not effective_dwh:
            # Tenter de recuperer le premier DWH actif
            try:
                dwh_list = execute_query(
                    "SELECT TOP 1 code FROM APP_DWH WHERE actif = 1 ORDER BY id",
                    use_cache=True,
                )
                if dwh_list:
                    effective_dwh = dwh_list[0]["code"]
            except Exception:
                pass

        result = _execute_builder_logic(config, annee, effective_dwh)
        return result

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"execute_excel_builder({builder_id}): {e}")
        return {"success": False, "error": str(e), "columns": MOIS_LABELS, "sections": []}


# =============================================================================
# ENDPOINT 7 — Exporter en fichier Excel (.xlsx)
# =============================================================================

@router.get("/{builder_id}/export")
async def export_excel_builder(
    builder_id: int,
    year: Optional[int] = None,
    dwh_code: Optional[str] = Header(None, alias="X-DWH-Code"),
):
    """
    Genere un fichier Excel (.xlsx) a partir du builder et le retourne en telechargement.
    Le parametre 'year' peut etre passe en query string (?year=2025).
    """
    from datetime import date
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    try:
        # Recuperer la config du builder
        results = _eb_read(
            "SELECT * FROM APP_ExcelBuilders WHERE id = ?",
            (builder_id,),
            dwh_code=dwh_code,
        )
        if not results:
            raise HTTPException(status_code=404, detail="Excel Builder non trouve")

        builder = results[0]
        builder_name = builder.get("name") or f"Excel Builder {builder_id}"
        raw_config = builder.get("config") or "{}"
        config = json.loads(raw_config) if isinstance(raw_config, str) else raw_config

        annee = year or date.today().year
        effective_dwh = dwh_code or _ctx_dwh.get()

        if not effective_dwh:
            try:
                dwh_list = execute_query(
                    "SELECT TOP 1 code FROM APP_DWH WHERE actif = 1 ORDER BY id",
                    use_cache=True,
                )
                if dwh_list:
                    effective_dwh = dwh_list[0]["code"]
            except Exception:
                pass

        # Executer la logique pour obtenir les donnees
        data = _execute_builder_logic(config, annee, effective_dwh)
        colonnes = data.get("columns", MOIS_LABELS)
        sections = data.get("sections", [])

        # ── Construction du classeur Excel ────────────────────────────────────

        wb = Workbook()
        ws = wb.active
        ws.title = f"{builder_name[:28]}"

        # Couleurs et styles de base
        BLEU_TITRE    = "2563EB"
        BLEU_FONCE    = "1E40AF"
        BLANC         = "FFFFFF"
        GRIS_CLAIR    = "F3F4F6"
        GRIS_MOYEN    = "E5E7EB"
        NOIR          = "111827"
        VERT_TOTAL    = "D1FAE5"
        VERT_FONCE    = "065F46"

        def _fill(hex_color: str) -> PatternFill:
            return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

        def _font(bold: bool = False, color: str = NOIR, size: int = 11) -> Font:
            return Font(bold=bold, color=color, size=size, name="Calibri")

        def _border_thin() -> Border:
            thin = Side(style="thin", color="D1D5DB")
            return Border(left=thin, right=thin, top=thin, bottom=thin)

        def _align(horizontal: str = "left", vertical: str = "center") -> Alignment:
            return Alignment(horizontal=horizontal, vertical=vertical, wrap_text=False)

        # ── Largeur des colonnes ──────────────────────────────────────────────
        # Colonne A : label  |  Colonnes B..M : mois  |  Colonne N : Total
        nb_mois = len(colonnes)
        col_label_idx = 1          # A
        col_total_idx = nb_mois + 2  # apres les 12 mois + 1 colonne label

        ws.column_dimensions["A"].width = 36
        for m in range(1, nb_mois + 1):
            ws.column_dimensions[get_column_letter(m + 1)].width = 12
        ws.column_dimensions[get_column_letter(col_total_idx)].width = 14

        ligne_courante = 1

        # ── Ligne de titre (fusionnee, fond bleu fonce) ───────────────────────
        titre_texte = f"{builder_name}  —  Exercice {annee}"
        derniere_col_lettre = get_column_letter(col_total_idx)

        ws.merge_cells(f"A{ligne_courante}:{derniere_col_lettre}{ligne_courante}")
        cellule_titre = ws.cell(row=ligne_courante, column=1, value=titre_texte)
        cellule_titre.fill = _fill(BLEU_FONCE)
        cellule_titre.font = _font(bold=True, color=BLANC, size=14)
        cellule_titre.alignment = _align(horizontal="center")
        cellule_titre.border = _border_thin()
        ws.row_dimensions[ligne_courante].height = 28
        ligne_courante += 1

        # ── Ligne d'en-tetes des colonnes ─────────────────────────────────────
        entetes = ["Libelle"] + colonnes + ["Total"]
        for col_idx, entete in enumerate(entetes, start=1):
            cell = ws.cell(row=ligne_courante, column=col_idx, value=entete)
            cell.fill = _fill(BLEU_TITRE)
            cell.font = _font(bold=True, color=BLANC, size=10)
            cell.alignment = _align(horizontal="center" if col_idx > 1 else "left")
            cell.border = _border_thin()
        ws.row_dimensions[ligne_courante].height = 20
        ligne_courante += 1

        # ── Donnees par section ───────────────────────────────────────────────
        for section in sections:
            section_label = section.get("label", "")
            section_color_hex = (section.get("color") or "#2563EB").lstrip("#")
            # S'assurer que la couleur est valide (6 hex chars)
            if len(section_color_hex) != 6:
                section_color_hex = "2563EB"

            section_rows = section.get("rows", [])

            # En-tete de section (ligne fusionnee, couleur de la section)
            ws.merge_cells(f"A{ligne_courante}:{derniere_col_lettre}{ligne_courante}")
            cell_sec = ws.cell(row=ligne_courante, column=1, value=section_label)
            cell_sec.fill = _fill(section_color_hex)
            cell_sec.font = _font(bold=True, color=BLANC, size=11)
            cell_sec.alignment = _align(horizontal="left")
            cell_sec.border = _border_thin()
            ws.row_dimensions[ligne_courante].height = 18
            ligne_courante += 1

            for row_data in section_rows:
                row_label = row_data.get("label", "")
                row_type  = row_data.get("type", "data")
                valeurs   = row_data.get("values", [0.0] * nb_mois)
                total_val = row_data.get("total", 0.0)

                est_total   = row_type in ("total", "formula")

                # Fond alterne : gris clair pour les lignes data
                if est_total:
                    fond = _fill(VERT_TOTAL)
                    font_row = _font(bold=True, color=VERT_FONCE, size=10)
                else:
                    fond = _fill(GRIS_CLAIR)
                    font_row = _font(bold=False, color=NOIR, size=10)

                # Colonne label
                cell_lbl = ws.cell(row=ligne_courante, column=1, value=row_label)
                cell_lbl.fill = fond
                cell_lbl.font = font_row
                cell_lbl.alignment = _align(horizontal="left")
                cell_lbl.border = _border_thin()

                # Colonnes mois
                for m_idx, val in enumerate(valeurs[:nb_mois], start=2):
                    cell_val = ws.cell(row=ligne_courante, column=m_idx, value=round(val, 2))
                    cell_val.fill = fond
                    cell_val.font = font_row
                    cell_val.alignment = _align(horizontal="right")
                    cell_val.border = _border_thin()
                    cell_val.number_format = "#,##0.00"

                # Colonne Total
                cell_tot = ws.cell(row=ligne_courante, column=col_total_idx, value=round(total_val, 2))
                cell_tot.fill = _fill(GRIS_MOYEN) if not est_total else _fill(VERT_TOTAL)
                cell_tot.font = _font(bold=True, color=VERT_FONCE if est_total else NOIR, size=10)
                cell_tot.alignment = _align(horizontal="right")
                cell_tot.border = _border_thin()
                cell_tot.number_format = "#,##0.00"

                ws.row_dimensions[ligne_courante].height = 16
                ligne_courante += 1

            # Ligne de separation vide entre sections
            ws.row_dimensions[ligne_courante].height = 6
            ligne_courante += 1

        # Figer la premiere ligne de donnees (titre + en-tetes = 2 lignes)
        ws.freeze_panes = ws.cell(row=3, column=2)

        # ── Sauvegarder dans un fichier temporaire ────────────────────────────
        nom_fichier = f"excel_builder_{builder_id}_{annee}.xlsx"
        tmp_dir = tempfile.gettempdir()
        chemin_tmp = os.path.join(tmp_dir, nom_fichier)

        wb.save(chemin_tmp)

        return FileResponse(
            path=chemin_tmp,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=nom_fichier,
            background=None,
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"export_excel_builder({builder_id}): {e}")
        raise HTTPException(status_code=500, detail=str(e))
