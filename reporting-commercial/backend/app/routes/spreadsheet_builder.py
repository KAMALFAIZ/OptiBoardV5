"""
Spreadsheet Builder — API Router
=================================
Module classeur avec FortuneSheet (affichage données + formules utilisateur) :
- CRUD complet des configurations spreadsheet
- Chargement de données depuis DataSources existants
- Multi-feuilles (chaque feuille liée à une datasource)
- Sauvegarde de l'état utilisateur (formules, formatage)
- Export Excel
"""

import io
import json
import logging
import re
import time
import hashlib
from datetime import datetime
from typing import Dict, Any, List, Optional

from fastapi import APIRouter, HTTPException, Header, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from ..database_unified import (
    execute_central as execute_query,
    central_cursor as get_db_cursor,
    DWHConnectionManager,
    execute_client,
    client_manager,
)
from ..services.datasource_resolver import datasource_resolver
from ..services.parameter_resolver import inject_params

logger = logging.getLogger("SpreadsheetBuilder")

router = APIRouter(prefix="/api/spreadsheet", tags=["Spreadsheet Builder"])

_tables_initialized = False


# =============================================================================
# DB READ HELPER (même pattern que pivot_v2 / gridview_builder)
# =============================================================================

def _ss_read(query: str, params: tuple = (), dwh_code: str = None) -> list:
    if dwh_code:
        try:
            if client_manager.has_client_db(dwh_code):
                if "WHERE id = ?" in query and len(params) == 1:
                    central = execute_query(
                        "SELECT code FROM APP_Spreadsheets WHERE id = ?", params, use_cache=False
                    )
                    if central and central[0].get("code"):
                        code = central[0]["code"]
                        code_query = query.replace("WHERE id = ?", "WHERE code = ?")
                        try:
                            result = execute_client(code_query, (code,), dwh_code=dwh_code, use_cache=False)
                            if result:
                                return result
                        except Exception:
                            pass
                result = execute_client(query, params, dwh_code=dwh_code, use_cache=False)
                if result:
                    return result
        except Exception as e:
            logger.debug(f"_ss_read client fallback ({dwh_code}): {e}")
    return execute_query(query, params, use_cache=False)


def _generate_code(nom: str) -> str:
    slug = re.sub(r'[^a-z0-9]+', '_', nom.lower().strip())[:40].strip('_')
    suffix = hashlib.md5(f"{nom}{time.time()}".encode()).hexdigest()[:4]
    return f"SS_{slug}_{suffix}"


# =============================================================================
# TABLE INIT
# =============================================================================

def init_spreadsheet_tables():
    global _tables_initialized
    if _tables_initialized:
        return True
    try:
        with get_db_cursor() as cursor:
            cursor.execute("""
                IF NOT EXISTS (SELECT * FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_NAME = 'APP_Spreadsheets')
                BEGIN
                    CREATE TABLE APP_Spreadsheets (
                        id INT IDENTITY(1,1) PRIMARY KEY,
                        nom NVARCHAR(200) NOT NULL,
                        code VARCHAR(100),
                        description NVARCHAR(500),
                        sheets_config NVARCHAR(MAX),
                        features NVARCHAR(MAX),
                        application NVARCHAR(100),
                        is_public BIT DEFAULT 0,
                        created_by INT,
                        actif BIT DEFAULT 1,
                        created_at DATETIME DEFAULT GETDATE(),
                        updated_at DATETIME DEFAULT GETDATE()
                    )
                END
            """)

            cursor.execute("""
                IF NOT EXISTS (SELECT * FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_NAME = 'APP_Spreadsheet_User_State')
                BEGIN
                    CREATE TABLE APP_Spreadsheet_User_State (
                        id INT IDENTITY(1,1) PRIMARY KEY,
                        spreadsheet_id INT NOT NULL,
                        user_id INT NOT NULL,
                        sheet_data NVARCHAR(MAX),
                        updated_at DATETIME DEFAULT GETDATE(),
                        CONSTRAINT FK_SSUserState_SS FOREIGN KEY (spreadsheet_id)
                            REFERENCES APP_Spreadsheets(id) ON DELETE CASCADE,
                        CONSTRAINT UQ_SSUserState UNIQUE (spreadsheet_id, user_id)
                    )
                END
            """)

            new_columns = [
                ("code", "VARCHAR(100)"),
                ("application", "NVARCHAR(100)"),
                ("actif", "BIT DEFAULT 1"),
                ("nom", "NVARCHAR(200)"),
                ("sheets_config", "NVARCHAR(MAX)"),
                ("features", "NVARCHAR(MAX)"),
                ("is_public", "BIT DEFAULT 0"),
            ]
            for col_name, col_def in new_columns:
                try:
                    cursor.execute(f"""
                        IF NOT EXISTS (SELECT 1 FROM INFORMATION_SCHEMA.COLUMNS
                                       WHERE TABLE_NAME='APP_Spreadsheets' AND COLUMN_NAME='{col_name}')
                        ALTER TABLE APP_Spreadsheets ADD {col_name} {col_def}
                    """)
                except Exception:
                    pass

        _tables_initialized = True
        logger.info("Spreadsheet tables initialized")
        return True
    except Exception as e:
        logger.error(f"Error initializing Spreadsheet tables: {e}")
        return False


# =============================================================================
# PYDANTIC SCHEMAS
# =============================================================================

class SheetDef(BaseModel):
    name: str
    data_source_code: Optional[str] = None
    data_source_id: Optional[int] = None
    column_mapping: Optional[List[Dict[str, Any]]] = None
    options: Optional[Dict[str, Any]] = None
    imported_celldata: Optional[List[Dict[str, Any]]] = None
    imported_config: Optional[Dict[str, Any]] = None
    imported_row_count: Optional[int] = None
    imported_column_count: Optional[int] = None

class SpreadsheetCreate(BaseModel):
    nom: str
    description: Optional[str] = None
    sheets: List[SheetDef]
    features: Optional[Dict[str, Any]] = None
    application: Optional[str] = None
    is_public: Optional[bool] = False

class SpreadsheetUpdate(BaseModel):
    nom: Optional[str] = None
    description: Optional[str] = None
    sheets: Optional[List[SheetDef]] = None
    features: Optional[Dict[str, Any]] = None
    application: Optional[str] = None
    is_public: Optional[bool] = None

class SheetDataRequest(BaseModel):
    context: Optional[Dict[str, Any]] = {}
    sheet_index: Optional[int] = None


# =============================================================================
# CRUD ENDPOINTS
# =============================================================================

@router.get("/sheets")
async def list_spreadsheets(
    user_id: Optional[int] = Query(None),
    dwh_code: Optional[str] = Header(None, alias="X-DWH-Code"),
):
    try:
        init_spreadsheet_tables()
        rows = _ss_read(
            "SELECT id, COALESCE(nom, name) AS nom, code, description, application, is_public, created_by, created_at, updated_at "
            "FROM APP_Spreadsheets WHERE actif = 1 ORDER BY updated_at DESC",
            dwh_code=dwh_code,
        )
        return {"success": True, "data": rows}
    except Exception as e:
        logger.error(f"list_spreadsheets: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/sheets/{sheet_id}")
async def get_spreadsheet(
    sheet_id: int,
    dwh_code: Optional[str] = Header(None, alias="X-DWH-Code"),
):
    try:
        init_spreadsheet_tables()
        rows = _ss_read("SELECT * FROM APP_Spreadsheets WHERE id = ?", (sheet_id,), dwh_code=dwh_code)
        if not rows:
            raise HTTPException(status_code=404, detail="Spreadsheet non trouve")
        row = rows[0]
        row['sheets_config'] = json.loads(row.get('sheets_config') or '[]')
        row['features'] = json.loads(row.get('features') or '{}')
        return {"success": True, "data": row}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"get_spreadsheet: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/sheets")
async def create_spreadsheet(
    payload: SpreadsheetCreate,
    dwh_code: Optional[str] = Header(None, alias="X-DWH-Code"),
    user_id_hdr: Optional[str] = Header(None, alias="X-User-Id"),
):
    try:
        if not (payload.nom or '').strip():
            raise HTTPException(status_code=422, detail="Le nom du classeur est requis")
        if not payload.sheets or len(payload.sheets) == 0:
            raise HTTPException(status_code=422, detail="Au moins une feuille est requise")
        init_spreadsheet_tables()
        code = _generate_code(payload.nom)
        uid = int(user_id_hdr) if user_id_hdr else None
        sheets_json = json.dumps([s.dict() for s in payload.sheets], ensure_ascii=False)
        features_json = json.dumps(payload.features or {}, ensure_ascii=False)

        with get_db_cursor() as cursor:
            cursor.execute("""
                INSERT INTO APP_Spreadsheets
                (nom, name, code, description, sheets_config, features, application, is_public, created_by, actif, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 1, GETDATE(), GETDATE())
            """, (
                payload.nom, payload.nom, code, payload.description,
                sheets_json, features_json,
                payload.application, 1 if payload.is_public else 0, uid,
            ))
            cursor.execute("SELECT SCOPE_IDENTITY() AS id")
            row = cursor.fetchone()
            new_id = row[0] if row and row[0] is not None else None

        if new_id is None:
            result = execute_query("SELECT TOP 1 id FROM APP_Spreadsheets WHERE code = ? ORDER BY id DESC", (code,), use_cache=False)
            new_id = result[0]['id'] if result else 0

        return {"success": True, "id": int(new_id), "code": code}
    except Exception as e:
        logger.error(f"create_spreadsheet: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/sheets/{sheet_id}")
async def update_spreadsheet(
    sheet_id: int,
    payload: SpreadsheetUpdate,
    dwh_code: Optional[str] = Header(None, alias="X-DWH-Code"),
    user_id_hdr: Optional[str] = Header(None, alias="X-User-Id"),
):
    try:
        init_spreadsheet_tables()
        uid = int(user_id_hdr) if user_id_hdr else None
        if uid:
            with get_db_cursor() as cursor:
                cursor.execute("SELECT created_by FROM APP_Spreadsheets WHERE id = ?", (sheet_id,))
                row = cursor.fetchone()
                if row and row[0] and int(row[0]) != uid:
                    raise HTTPException(status_code=403, detail="Vous ne pouvez modifier que vos propres classeurs")
        sets, params = [], []
        if payload.nom is not None:
            sets.append("nom = ?"); params.append(payload.nom)
            sets.append("name = ?"); params.append(payload.nom)
        if payload.description is not None:
            sets.append("description = ?"); params.append(payload.description)
        if payload.sheets is not None:
            sets.append("sheets_config = ?")
            params.append(json.dumps([s.dict() for s in payload.sheets], ensure_ascii=False))
        if payload.features is not None:
            sets.append("features = ?")
            params.append(json.dumps(payload.features, ensure_ascii=False))
        if payload.application is not None:
            sets.append("application = ?"); params.append(payload.application)
        if payload.is_public is not None:
            sets.append("is_public = ?"); params.append(1 if payload.is_public else 0)

        if not sets:
            return {"success": True, "message": "Aucune modification"}

        sets.append("updated_at = GETDATE()")
        params.append(sheet_id)
        with get_db_cursor() as cursor:
            cursor.execute(f"UPDATE APP_Spreadsheets SET {', '.join(sets)} WHERE id = ?", tuple(params))

        return {"success": True}
    except Exception as e:
        logger.error(f"update_spreadsheet: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/sheets/{sheet_id}")
async def delete_spreadsheet(
    sheet_id: int,
    dwh_code: Optional[str] = Header(None, alias="X-DWH-Code"),
    user_id_hdr: Optional[str] = Header(None, alias="X-User-Id"),
):
    try:
        uid = int(user_id_hdr) if user_id_hdr else None
        with get_db_cursor() as cursor:
            if uid:
                cursor.execute("SELECT created_by FROM APP_Spreadsheets WHERE id = ?", (sheet_id,))
                row = cursor.fetchone()
                if row and row[0] and int(row[0]) != uid:
                    raise HTTPException(status_code=403, detail="Vous ne pouvez supprimer que vos propres classeurs")
            cursor.execute("UPDATE APP_Spreadsheets SET actif = 0 WHERE id = ?", (sheet_id,))
        return {"success": True}
    except Exception as e:
        logger.error(f"delete_spreadsheet: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# =============================================================================
# DATA LOADING — Charge les données d'une datasource pour une feuille
# =============================================================================

@router.post("/sheets/{sheet_id}/data")
async def get_sheet_data(
    sheet_id: int,
    request: SheetDataRequest = SheetDataRequest(),
    dwh_code: Optional[str] = Header(None, alias="X-DWH-Code"),
    user_id_hdr: Optional[str] = Header(None, alias="X-User-Id"),
):
    try:
        init_spreadsheet_tables()
        rows = _ss_read("SELECT * FROM APP_Spreadsheets WHERE id = ?", (sheet_id,), dwh_code=dwh_code)
        if not rows:
            raise HTTPException(status_code=404, detail="Spreadsheet non trouve")

        config = rows[0]
        sheets_config = json.loads(config.get('sheets_config') or '[]')
        context = request.context or {}

        result_sheets = []

        target_indices = range(len(sheets_config))
        if request.sheet_index is not None:
            target_indices = [request.sheet_index]

        for idx in target_indices:
            if idx >= len(sheets_config):
                continue
            sheet_def = sheets_config[idx]
            ds_code = sheet_def.get('data_source_code')
            ds_id = sheet_def.get('data_source_id')
            col_mapping = sheet_def.get('column_mapping') or []
            sheet_name = sheet_def.get('name', f'Feuille {idx + 1}')

            if not ds_code and not ds_id:
                imported = sheet_def.get('imported_celldata')
                if imported:
                    result_sheets.append({
                        "name": sheet_name,
                        "index": idx,
                        "celldata": imported,
                        "headers": [],
                        "row_count": sheet_def.get('imported_row_count', 0),
                        "column_count": sheet_def.get('imported_column_count', 0),
                        "config": sheet_def.get('imported_config', {}),
                    })
                else:
                    result_sheets.append({
                        "name": sheet_name,
                        "index": idx,
                        "celldata": [],
                        "headers": [],
                        "row_count": 0,
                    })
                continue

            base_query = None
            try:
                if ds_code:
                    ds = datasource_resolver.resolve_by_code(ds_code, dwh_code)
                    base_query = ds.query_template
                elif ds_id:
                    ds = datasource_resolver.resolve_by_id(ds_id, dwh_code)
                    base_query = ds.query_template
            except (ValueError, Exception) as e:
                if ds_id:
                    source = execute_query(
                        "SELECT * FROM APP_DataSources WHERE id = ?", (ds_id,), use_cache=False
                    )
                    if source:
                        base_query = source[0]['query_template']

            if not base_query:
                result_sheets.append({
                    "name": sheet_name,
                    "index": idx,
                    "celldata": [],
                    "headers": [],
                    "row_count": 0,
                    "error": "Source de donnees non trouvee",
                })
                continue

            if '@societe_filter' in base_query:
                societe_val = context.get('societe') or context.get('societe_filter')
                if societe_val:
                    base_query = base_query.replace('@societe_filter', f"societe = '{societe_val}'")
                else:
                    base_query = base_query.replace('@societe_filter', '1=1')

            final_query = inject_params(base_query, context)

            effective_dwh = dwh_code
            if not effective_dwh:
                try:
                    dwh_list = execute_query(
                        "SELECT TOP 1 code FROM APP_DWH WHERE actif = 1 ORDER BY id", use_cache=True
                    )
                    if dwh_list:
                        effective_dwh = dwh_list[0]['code']
                except Exception:
                    pass

            data_rows = []
            if effective_dwh:
                try:
                    data_rows = DWHConnectionManager.execute_dwh_query(
                        effective_dwh, final_query, use_cache=False
                    )
                except Exception as dwh_err:
                    logger.error(f"[SHEET {sheet_id}] DWH error: {dwh_err}")
                    result_sheets.append({
                        "name": sheet_name,
                        "index": idx,
                        "celldata": [],
                        "headers": [],
                        "row_count": 0,
                        "error": str(dwh_err)[:200],
                    })
                    continue
            else:
                data_rows = execute_query(final_query, use_cache=False)

            if not data_rows:
                result_sheets.append({
                    "name": sheet_name,
                    "index": idx,
                    "celldata": [],
                    "headers": [],
                    "row_count": 0,
                })
                continue

            headers = list(data_rows[0].keys())

            if col_mapping:
                mapped_headers = [m.get('label', m.get('field', '')) for m in col_mapping]
                mapped_fields = [m.get('field', '') for m in col_mapping]
            else:
                mapped_headers = headers
                mapped_fields = headers

            celldata = []
            for c_idx, h in enumerate(mapped_headers):
                celldata.append({"r": 0, "c": c_idx, "v": {"v": h, "m": h, "bl": 1, "ct": {"fa": "General", "t": "g"}}})

            for r_idx, row in enumerate(data_rows):
                for c_idx, field in enumerate(mapped_fields):
                    val = row.get(field)
                    if val is None:
                        continue
                    cell = {"r": r_idx + 1, "c": c_idx, "v": _format_cell_value(val)}
                    celldata.append(cell)

            numeric_cols = []
            for h in mapped_fields:
                sample = next((r.get(h) for r in data_rows[:10] if r.get(h) is not None), None)
                if isinstance(sample, (int, float)):
                    numeric_cols.append(h)
                else:
                    from decimal import Decimal
                    if isinstance(sample, Decimal):
                        numeric_cols.append(h)

            stats = {}
            for nc in numeric_cols:
                vals = [float(r.get(nc, 0) or 0) for r in data_rows if r.get(nc) is not None]
                if vals:
                    label = nc
                    for m in (col_mapping or []):
                        if m.get('field') == nc:
                            label = m.get('label', nc)
                            break
                    stats[nc] = {
                        "label": label,
                        "sum": round(sum(vals), 2),
                        "avg": round(sum(vals) / len(vals), 2),
                        "min": round(min(vals), 2),
                        "max": round(max(vals), 2),
                        "count": len(vals),
                    }

            chart_data = []
            text_col = None
            main_num_col = None
            for h in mapped_fields:
                if h not in numeric_cols and text_col is None:
                    text_col = h
                if h in numeric_cols and main_num_col is None:
                    main_num_col = h
            if text_col and main_num_col:
                for row in data_rows[:20]:
                    label = str(row.get(text_col, ''))[:25]
                    entry = {"name": label}
                    for nc in numeric_cols[:4]:
                        entry[nc] = float(row.get(nc, 0) or 0)
                    chart_data.append(entry)

            result_sheets.append({
                "name": sheet_name,
                "index": idx,
                "celldata": celldata,
                "headers": mapped_headers,
                "fields": mapped_fields,
                "row_count": len(data_rows),
                "column_count": len(mapped_headers),
                "stats": stats,
                "chart_data": chart_data,
                "numeric_cols": numeric_cols,
            })

        return {"success": True, "sheets": result_sheets}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"get_sheet_data: {e}")
        raise HTTPException(status_code=500, detail=str(e))


def _format_cell_value(val) -> dict:
    from decimal import Decimal
    if isinstance(val, (int, float)):
        return {"v": val, "m": str(val), "ct": {"fa": "General", "t": "n"}}
    if isinstance(val, Decimal):
        fv = float(val)
        return {"v": fv, "m": str(fv), "ct": {"fa": "General", "t": "n"}}
    if isinstance(val, datetime):
        return {"v": val.strftime("%Y-%m-%d"), "m": val.strftime("%d/%m/%Y"), "ct": {"fa": "yyyy-mm-dd", "t": "d"}}
    from datetime import date
    if isinstance(val, date):
        return {"v": val.strftime("%Y-%m-%d"), "m": val.strftime("%d/%m/%Y"), "ct": {"fa": "yyyy-mm-dd", "t": "d"}}
    sv = str(val)
    return {"v": sv, "m": sv, "ct": {"fa": "General", "t": "g"}}


# =============================================================================
# USER STATE (preferences utilisateur — même pattern que pivot_v2)
# =============================================================================

@router.get("/sheets/{sheet_id}/state/{user_id}")
async def get_user_state(
    sheet_id: int,
    user_id: int,
    dwh_code: Optional[str] = Header(None, alias="X-DWH-Code"),
):
    try:
        init_spreadsheet_tables()
        rows = execute_query(
            "SELECT * FROM APP_Spreadsheet_User_State WHERE spreadsheet_id = ? AND user_id = ?",
            (sheet_id, user_id), use_cache=False,
        )
        if not rows:
            return {"success": True, "data": None}
        state = rows[0]
        state['sheet_data'] = json.loads(state.get('sheet_data') or 'null')
        return {"success": True, "data": state}
    except Exception as e:
        logger.error(f"get_user_state: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/sheets/{sheet_id}/state/{user_id}")
async def save_user_state(
    sheet_id: int,
    user_id: int,
    payload: dict,
    dwh_code: Optional[str] = Header(None, alias="X-DWH-Code"),
):
    try:
        init_spreadsheet_tables()
        sheet_data_json = json.dumps(payload.get('sheet_data', {}), ensure_ascii=False)
        with get_db_cursor() as cursor:
            cursor.execute(
                "MERGE APP_Spreadsheet_User_State AS target "
                "USING (SELECT ? AS spreadsheet_id, ? AS user_id) AS source "
                "ON target.spreadsheet_id = source.spreadsheet_id AND target.user_id = source.user_id "
                "WHEN MATCHED THEN UPDATE SET sheet_data = ?, updated_at = GETDATE() "
                "WHEN NOT MATCHED THEN INSERT (spreadsheet_id, user_id, sheet_data, updated_at) "
                "VALUES (?, ?, ?, GETDATE());",
                (sheet_id, user_id, sheet_data_json, sheet_id, user_id, sheet_data_json),
            )
        return {"success": True}
    except Exception as e:
        logger.error(f"save_user_state: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/sheets/{sheet_id}/state/{user_id}")
async def reset_user_state(
    sheet_id: int,
    user_id: int,
    dwh_code: Optional[str] = Header(None, alias="X-DWH-Code"),
):
    try:
        with get_db_cursor() as cursor:
            cursor.execute(
                "DELETE FROM APP_Spreadsheet_User_State WHERE spreadsheet_id = ? AND user_id = ?",
                (sheet_id, user_id),
            )
        return {"success": True}
    except Exception as e:
        logger.error(f"reset_user_state: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# =============================================================================
# EXPORT EXCEL
# =============================================================================

@router.post("/sheets/{sheet_id}/export")
async def export_spreadsheet(
    sheet_id: int,
    request: SheetDataRequest = SheetDataRequest(),
    dwh_code: Optional[str] = Header(None, alias="X-DWH-Code"),
):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        data_response = await get_sheet_data(sheet_id, request, dwh_code)
        if not data_response.get("success"):
            raise HTTPException(status_code=500, detail="Erreur chargement donnees")

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        for sheet_info in data_response["sheets"]:
            ws = wb.create_sheet(title=sheet_info["name"][:31])
            headers = sheet_info.get("headers", [])
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

            for c_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=c_idx, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            for cell_item in sheet_info.get("celldata", []):
                r = cell_item.get("r", 0)
                c = cell_item.get("c", 0)
                v = cell_item.get("v", {})
                if r == 0:
                    continue
                ws.cell(row=r + 1, column=c + 1, value=v.get("v", ""))

            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        max_len = max(max_len, len(str(cell.value or "")))
                    except Exception:
                        pass
                ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        rows_conf = _ss_read("SELECT nom FROM APP_Spreadsheets WHERE id = ?", (sheet_id,))
        filename = rows_conf[0]['nom'] if rows_conf else f"spreadsheet_{sheet_id}"
        safe_name = re.sub(r'[^\w\s\-]', '', filename)[:50]

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{safe_name}.xlsx"'},
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"export_spreadsheet: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# =============================================================================
# IMPORT EXCEL — Parse un fichier .xlsx et retourne du celldata FortuneSheet
# =============================================================================

@router.post("/import-excel")
async def import_excel(file: UploadFile = File(...)):
    try:
        import openpyxl

        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="Format non supporte. Utilisez un fichier .xlsx")

        contents = await file.read()
        if len(contents) > 100 * 1024 * 1024:
            raise HTTPException(status_code=400, detail="Fichier trop volumineux (max 100 Mo)")

        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        result_sheets = []

        for ws_idx, ws in enumerate(wb.worksheets):
            celldata = []
            max_row = 0
            max_col = 0

            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    r, c = cell.row - 1, cell.column - 1
                    max_row = max(max_row, r)
                    max_col = max(max_col, c)

                    cv = _format_cell_value(cell.value)

                    if cell.font and cell.font.bold:
                        cv["bl"] = 1
                    if cell.font and cell.font.italic:
                        cv["it"] = 1

                    if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb and cell.fill.fgColor.rgb != '00000000':
                        rgb = cell.fill.fgColor.rgb
                        if isinstance(rgb, str) and len(rgb) >= 6:
                            cv["bg"] = f"#{rgb[-6:]}"

                    if cell.font and cell.font.color and cell.font.color.rgb:
                        fc_rgb = cell.font.color.rgb
                        if isinstance(fc_rgb, str) and len(fc_rgb) >= 6:
                            cv["fc"] = f"#{fc_rgb[-6:]}"

                    celldata.append({"r": r, "c": c, "v": cv})

            merged = []
            for mr in ws.merged_cells.ranges:
                merged.append({
                    "r": mr.min_row - 1,
                    "c": mr.min_col - 1,
                    "rs": mr.max_row - mr.min_row + 1,
                    "cs": mr.max_col - mr.min_col + 1,
                })

            col_widths = {}
            for col_letter, dim in ws.column_dimensions.items():
                if dim.width and dim.width != 8.43:
                    col_idx = openpyxl.utils.column_index_from_string(col_letter) - 1
                    col_widths[str(col_idx)] = round(dim.width * 7.5)

            row_heights = {}
            for row_num, dim in ws.row_dimensions.items():
                if dim.height and dim.height != 15:
                    row_heights[str(row_num - 1)] = round(dim.height)

            sheet_config = {}
            if col_widths:
                sheet_config["columnlen"] = col_widths
            if row_heights:
                sheet_config["rowlen"] = row_heights
            if merged:
                sheet_config["merge"] = {
                    f"{m['r']}_{m['c']}": m for m in merged
                }

            result_sheets.append({
                "name": ws.title or f"Feuille {ws_idx + 1}",
                "celldata": celldata,
                "row_count": max_row + 1,
                "column_count": max_col + 1,
                "config": sheet_config,
            })

        wb.close()
        return {"success": True, "sheets": result_sheets, "filename": file.filename}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"import_excel: {e}")
        raise HTTPException(status_code=500, detail=str(e))
