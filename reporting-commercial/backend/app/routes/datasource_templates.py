"""
Routes pour la gestion des DataSources Templates
================================================
CRUD pour les templates de datasources (centraux et overrides par DWH)
"""

import hashlib
import io
import logging
import re
import time
from datetime import datetime, date
from decimal import Decimal
from fastapi import APIRouter, HTTPException, Header, Query
from fastapi.responses import StreamingResponse
from typing import Optional, Dict, Any, List
from pydantic import BaseModel
from ..database_unified import execute_central as execute_query, central_cursor as get_db_cursor, DWHConnectionManager, dwh_cursor

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/datasources", tags=["DataSource Templates"])

# Cache TTL pour la détection de champs (10 minutes)
_UNIFIED_FIELDS_CACHE: Dict[str, Any] = {}
_UNIFIED_FIELDS_TTL = 600


# =============================================================================
# SCHEMAS
# =============================================================================

class DataSourceTemplateCreate(BaseModel):
    code: str
    nom: str
    type: str = "query"
    category: str = "Ventes"
    description: Optional[str] = None
    query_template: str
    parameters: Optional[str] = "[]"
    is_system: bool = False
    actif: bool = True


class DataSourceTemplateUpdate(BaseModel):
    nom: Optional[str] = None
    type: Optional[str] = None
    category: Optional[str] = None
    description: Optional[str] = None
    query_template: Optional[str] = None
    parameters: Optional[str] = None
    is_system: Optional[bool] = None
    actif: Optional[bool] = None


class DataSourceDrilldownRequest(BaseModel):
    filterField: Optional[str] = None
    filterValue: Optional[Any] = None
    context: Optional[Dict[str, Any]] = {}
    page: int = 1
    pageSize: int = 50
    sortField: Optional[str] = None
    sortDirection: Optional[str] = "asc"


# =============================================================================
# TEMPLATES (Base centrale)
# =============================================================================

@router.get("/templates")
async def get_templates(
    category: Optional[str] = Query(None, description="Filtrer par categorie"),
    search: Optional[str] = Query(None, description="Recherche par code ou nom")
):
    """Liste tous les templates de datasources depuis la base centrale"""
    try:
        query = """
            SELECT
                id, code, nom, type, category, description,
                query_template, parameters, is_system, actif, date_creation
            FROM APP_DataSources_Templates
            WHERE 1=1
        """
        params = []

        if category:
            query += " AND LOWER(category) = LOWER(?)"
            params.append(category)

        if search:
            query += " AND (code LIKE ? OR nom LIKE ?)"
            params.extend([f"%{search}%", f"%{search}%"])

        query += " ORDER BY category, code"

        templates = execute_query(query, tuple(params) if params else None, use_cache=False)

        # Compter par categorie
        categories_count = {}
        for t in templates:
            cat = t.get('category', 'Autre')
            categories_count[cat] = categories_count.get(cat, 0) + 1

        return {
            "success": True,
            "data": templates,
            "total": len(templates),
            "categories": categories_count
        }
    except Exception as e:
        logger.error(f"[ERROR] get_templates: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/templates/{template_id}")
async def get_template(template_id: int):
    """Recupere un template specifique par ID"""
    try:
        templates = execute_query(
            """
            SELECT
                id, code, nom, type, category, description,
                query_template, parameters, is_system, actif, date_creation
            FROM APP_DataSources_Templates
            WHERE id = ?
            """,
            (template_id,),
            use_cache=False
        )

        if not templates:
            raise HTTPException(status_code=404, detail="Template non trouve")

        return {
            "success": True,
            "data": templates[0]
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"[ERROR] get_template: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/templates/code/{code}")
async def get_template_by_code(code: str):
    """Recupere un template specifique par code"""
    try:
        templates = execute_query(
            """
            SELECT
                id, code, nom, type, category, description,
                query_template, parameters, is_system, actif, date_creation
            FROM APP_DataSources_Templates
            WHERE code = ?
            """,
            (code,),
            use_cache=False
        )

        if not templates:
            raise HTTPException(status_code=404, detail="Template non trouve")

        return {
            "success": True,
            "data": templates[0]
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"[ERROR] get_template_by_code: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/templates")
async def create_template(
    template: DataSourceTemplateCreate,
    x_user_role: str = Header(None, alias="X-User-Role"),
    x_user_id: str = Header(None, alias="X-User-Id")
):
    """Cree un nouveau template de datasource"""
    try:
        # Verification role pour templates systeme
        if template.is_system and x_user_role != "superadmin":
            raise HTTPException(status_code=403, detail="Seul un superadmin peut creer des templates systeme")

        # Verifier si le code existe deja
        existing = execute_query(
            "SELECT id FROM APP_DataSources_Templates WHERE code = ?",
            (template.code,),
            use_cache=False
        )
        if existing:
            raise HTTPException(status_code=400, detail=f"Le code '{template.code}' existe deja")

        # Inserer le template
        with get_db_cursor() as cursor:
            cursor.execute(
                """
                INSERT INTO APP_DataSources_Templates
                (code, nom, type, category, description, query_template, parameters, is_system, actif)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    template.code,
                    template.nom,
                    template.type,
                    template.category,
                    template.description,
                    template.query_template,
                    template.parameters,
                    1 if template.is_system else 0,
                    1 if template.actif else 0
                )
            )

            # Recuperer l'ID cree
            cursor.execute("SELECT SCOPE_IDENTITY() as id")
            result = cursor.fetchone()
            new_id = result[0] if result else None

        return {
            "success": True,
            "message": "Template cree avec succes",
            "id": new_id
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"[ERROR] create_template: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/templates/{template_id}")
async def update_template(
    template_id: int,
    template: DataSourceTemplateUpdate,
    x_user_role: str = Header(None, alias="X-User-Role"),
    x_user_id: str = Header(None, alias="X-User-Id")
):
    """Met a jour un template existant"""
    try:
        # Verifier que le template existe
        existing = execute_query(
            "SELECT id, is_system FROM APP_DataSources_Templates WHERE id = ?",
            (template_id,),
            use_cache=False
        )
        if not existing:
            raise HTTPException(status_code=404, detail="Template non trouve")

        # Verification role pour templates systeme
        if existing[0].get('is_system') and x_user_role != "superadmin":
            raise HTTPException(status_code=403, detail="Seul un superadmin peut modifier les templates systeme")

        # Construire la requete de mise a jour
        updates = []
        params = []

        if template.nom is not None:
            updates.append("nom = ?")
            params.append(template.nom)
        if template.type is not None:
            updates.append("type = ?")
            params.append(template.type)
        if template.category is not None:
            updates.append("category = ?")
            params.append(template.category)
        if template.description is not None:
            updates.append("description = ?")
            params.append(template.description)
        if template.query_template is not None:
            updates.append("query_template = ?")
            params.append(template.query_template)
        if template.parameters is not None:
            updates.append("parameters = ?")
            params.append(template.parameters)
        if template.is_system is not None and x_user_role == "superadmin":
            updates.append("is_system = ?")
            params.append(1 if template.is_system else 0)
        if template.actif is not None:
            updates.append("actif = ?")
            params.append(1 if template.actif else 0)

        if not updates:
            return {"success": True, "message": "Aucune modification"}

        params.append(template_id)

        with get_db_cursor() as cursor:
            cursor.execute(
                f"UPDATE APP_DataSources_Templates SET {', '.join(updates)} WHERE id = ?",
                tuple(params)
            )

        return {
            "success": True,
            "message": "Template mis a jour avec succes"
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"[ERROR] update_template: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/templates/id/{template_id}")
async def delete_template(
    template_id: int,
    x_user_role: str = Header(None, alias="X-User-Role")
):
    """Supprime un template"""
    try:
        # Verifier que le template existe
        existing = execute_query(
            "SELECT id, is_system, code FROM APP_DataSources_Templates WHERE id = ?",
            (template_id,),
            use_cache=False
        )
        if not existing:
            raise HTTPException(status_code=404, detail="Template non trouve")

        # Verification role pour templates systeme
        if existing[0].get('is_system') and x_user_role != "superadmin":
            raise HTTPException(status_code=403, detail="Seul un superadmin peut supprimer les templates systeme")

        with get_db_cursor() as cursor:
            cursor.execute("DELETE FROM APP_DataSources_Templates WHERE id = ?", (template_id,))

        return {
            "success": True,
            "message": f"Template '{existing[0].get('code')}' supprime avec succes"
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"[ERROR] delete_template: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# =============================================================================
# OVERRIDES (Base DWH client - pour future implementation)
# =============================================================================

@router.get("/overrides")
async def get_overrides(
    x_dwh_code: str = Header(None, alias="X-DWH-Code")
):
    """Liste les overrides de datasources pour un DWH specifique"""
    # Pour l'instant, retourne une liste vide car les overrides sont par DWH
    # Cette fonctionnalite sera implementee quand on aura le contexte DWH
    return {
        "success": True,
        "data": [],
        "message": "Les overrides sont specifiques par DWH client"
    }


# =============================================================================
# TEST DE REQUETE
# =============================================================================

# =============================================================================
# ROUTE UNIFIEE POUR LES BUILDERS (Templates + Sources locales)
# =============================================================================

@router.get("/unified")
async def get_unified_datasources(
    category: Optional[str] = Query(None, description="Filtrer par categorie"),
    search: Optional[str] = Query(None, description="Recherche par code ou nom"),
    include_local: bool = Query(True, description="Inclure les sources locales APP_DataSources")
):
    """
    Liste unifiee des DataSources pour les Builders (Dashboard, Pivot, GridView).

    Combine:
    1. APP_DataSources_Templates (templates centraux) - prioritaires
    2. APP_DataSources (sources locales) - si include_local=True

    Retourne une liste avec un champ 'origin' pour distinguer la source.
    """
    try:
        result = []
        seen_codes = set()

        # 1. D'abord les templates (prioritaires)
        template_query = """
            SELECT
                id, code, nom, type, category, description,
                query_template, parameters, is_system, actif,
                'template' as origin
            FROM APP_DataSources_Templates
            WHERE actif = 1
        """
        template_params = []

        if category:
            template_query += " AND LOWER(category) = LOWER(?)"
            template_params.append(category)

        if search:
            template_query += " AND (code LIKE ? OR nom LIKE ?)"
            template_params.extend([f"%{search}%", f"%{search}%"])

        template_query += " ORDER BY category, nom"

        templates = execute_query(template_query, tuple(template_params) if template_params else None, use_cache=False)

        for t in templates:
            t['origin'] = 'template'
            t['origin_label'] = 'Template'
            t['can_edit'] = False  # Les templates ne sont pas editables directement
            result.append(t)
            if t.get('code'):
                seen_codes.add(t['code'])

        # 2. Ensuite les sources locales (si pas deja un template avec le meme code)
        if include_local:
            local_query = """
                SELECT
                    id, nom, type, description,
                    query_template, parameters,
                    'local' as origin
                FROM APP_DataSources
            """
            local_params = []

            if search:
                local_query += " WHERE (nom LIKE ?)"
                local_params.append(f"%{search}%")

            local_query += " ORDER BY nom"

            local_sources = execute_query(local_query, tuple(local_params) if local_params else None, use_cache=False)

            for s in local_sources:
                # Generer un code si absent
                s['code'] = s.get('code') or f"LOCAL_{s['id']}"
                s['category'] = s.get('category') or 'custom'
                s['origin'] = 'local'
                s['origin_label'] = 'Source Locale'
                s['can_edit'] = True
                s['is_system'] = False
                s['actif'] = True

                # Ne pas ajouter si un template avec le meme code existe
                if s['code'] not in seen_codes:
                    result.append(s)

        # Grouper par categorie pour le frontend
        categories = {}
        for ds in result:
            cat = ds.get('category') or 'Autre'
            if cat not in categories:
                categories[cat] = []
            categories[cat].append(ds)

        return {
            "success": True,
            "data": result,
            "total": len(result),
            "categories": {k: len(v) for k, v in categories.items()},
            "grouped": categories
        }
    except Exception as e:
        logger.error(f"[ERROR] get_unified_datasources: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/unified/{identifier}")
async def get_unified_datasource(
    identifier: str,
    x_dwh_code: str = Header(None, alias="X-DWH-Code")
):
    """
    Recupere une DataSource par ID ou code.

    - Si identifier est numerique: cherche par ID (d'abord local, puis template)
    - Si identifier est une chaine: cherche par code (d'abord template, puis local)
    """
    from ..services.parameter_resolver import extract_parameters_from_query

    try:
        datasource = None
        origin = None

        # Determiner si c'est un ID ou un code
        is_numeric = identifier.isdigit()

        if is_numeric:
            ds_id = int(identifier)

            # Chercher d'abord dans les sources locales
            local = execute_query(
                "SELECT * FROM APP_DataSources WHERE id = ?",
                (ds_id,),
                use_cache=False
            )
            if local:
                datasource = local[0]
                datasource['origin'] = 'local'
                datasource['code'] = datasource.get('code') or f"LOCAL_{ds_id}"
                datasource['category'] = datasource.get('category') or 'custom'
            else:
                # Chercher dans les templates
                template = execute_query(
                    "SELECT * FROM APP_DataSources_Templates WHERE id = ?",
                    (ds_id,),
                    use_cache=False
                )
                if template:
                    datasource = template[0]
                    datasource['origin'] = 'template'
        else:
            # Chercher par code - d'abord templates (prioritaires)
            template = execute_query(
                "SELECT * FROM APP_DataSources_Templates WHERE code = ? AND actif = 1",
                (identifier,),
                use_cache=False
            )
            if template:
                datasource = template[0]
                datasource['origin'] = 'template'
            else:
                # Chercher dans les sources locales par nom
                local = execute_query(
                    "SELECT * FROM APP_DataSources WHERE nom = ?",
                    (identifier,),
                    use_cache=False
                )
                if local:
                    datasource = local[0]
                    datasource['origin'] = 'local'
                    datasource['code'] = datasource.get('code') or f"LOCAL_{datasource['id']}"

        if not datasource:
            raise HTTPException(status_code=404, detail=f"DataSource '{identifier}' non trouvee")

        # Parser les parametres si c'est une chaine JSON
        if datasource.get('parameters'):
            try:
                import json
                if isinstance(datasource['parameters'], str):
                    datasource['parameters'] = json.loads(datasource['parameters'])
            except:
                datasource['parameters'] = []
        else:
            datasource['parameters'] = []

        # Extraire les parametres de la requete si non definis
        if not datasource['parameters'] and datasource.get('query_template'):
            datasource['extracted_params'] = extract_parameters_from_query(datasource['query_template'])

        return {
            "success": True,
            "data": datasource
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"[ERROR] get_unified_datasource: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/unified/{identifier}/preview")
async def preview_unified_datasource(
    identifier: str,
    body: dict = {},
    x_dwh_code: str = Header(None, alias="X-DWH-Code")
):
    """
    Execute une DataSource (template ou locale) avec les parametres fournis.
    - limit=0 => pas de limite (viewer mode, toutes les donnees)
    - limit=100 (defaut) => mode builder/preview
    """
    from ..services.parameter_resolver import inject_params
    from ..database_unified import current_dwh_code as _ctx_dwh

    # Fallback: utiliser le ContextVar injecté par TenantContextMiddleware (DWH_CODE .env)
    effective_header_dwh = x_dwh_code or _ctx_dwh.get()

    try:
        context = body.get('context', {})
        limit = body.get('limit', 100)
        extra_filters = body.get('extra_filters', [])
        page = body.get('page', None)          # pagination serveur (1-based)
        page_size = body.get('page_size', 500)  # taille de page par defaut

        # Recuperer la datasource
        ds_response = await get_unified_datasource(identifier, effective_header_dwh)
        datasource = ds_response['data']

        query = datasource.get('query_template', '')
        if not query:
            raise HTTPException(status_code=400, detail="Aucune requete definie")

        # Injecter les parametres
        final_query = inject_params(query, context)

        # Appliquer extra_filters : WHERE [field] = 'value' sur le résultat
        if extra_filters:
            where_parts = []
            for ef in extra_filters:
                field = ef.get('field', '')
                value = ef.get('value')
                if field and value is not None:
                    safe_field = field.replace(']', ']]')
                    safe_value = str(value).replace("'", "''")
                    where_parts.append(f"[{safe_field}] = N'{safe_value}'")
            if where_parts:
                final_query = f"SELECT * FROM ({final_query}) AS __drilldown__ WHERE {' AND '.join(where_parts)}"

        # Pagination serveur : OFFSET/FETCH NEXT (SQL Server 2012+)
        # Si page est fourni, on utilise la pagination serveur au lieu de charger tout en mémoire
        count_query = None
        if page and page > 0 and page_size and page_size > 0:
            offset = (page - 1) * page_size
            # Compter le total pour la pagination
            count_query = f"SELECT COUNT(*) AS total FROM ({final_query}) AS __count__"
            # Ajouter ORDER BY si absent (requis pour OFFSET/FETCH)
            upper_q = final_query.upper()
            if "ORDER BY" not in upper_q:
                final_query = f"SELECT * FROM ({final_query}) AS __paged__ ORDER BY (SELECT NULL) OFFSET {offset} ROWS FETCH NEXT {page_size} ROWS ONLY"
            else:
                final_query = f"{final_query} OFFSET {offset} ROWS FETCH NEXT {page_size} ROWS ONLY"
        elif limit and limit > 0 and "SELECT" in final_query.upper() and "TOP" not in final_query.upper():
            # Mode classique : TOP N
            final_query = final_query.replace("SELECT", f"SELECT TOP {limit}", 1)

        # Executer sur le bon DWH
        effective_dwh = effective_header_dwh
        if not effective_dwh and datasource.get('origin') == 'template':
            try:
                dwh_list = execute_query(
                    "SELECT TOP 1 code FROM APP_DWH WHERE actif = 1 ORDER BY id",
                    use_cache=True
                )
                if dwh_list:
                    effective_dwh = dwh_list[0]['code']
            except Exception:
                pass

        exec_fn = (lambda q: DWHConnectionManager.execute_dwh_query(effective_dwh, q, use_cache=False)) if effective_dwh else (lambda q: execute_query(q, use_cache=False))

        # Recuperer le total si pagination
        total_count = None
        if count_query:
            try:
                count_result = exec_fn(count_query)
                if count_result:
                    total_count = count_result[0].get('total', 0)
            except Exception:
                pass

        results = exec_fn(final_query)

        # Extraire les colonnes avec types Python -> types grille
        import decimal as _dec_mod
        import datetime as _dt_mod

        def _py_to_col_type(v):
            if v is None:
                return None
            if isinstance(v, bool):
                return "boolean"
            if isinstance(v, (int, float, _dec_mod.Decimal)):
                return "number"
            if isinstance(v, (_dt_mod.datetime, _dt_mod.date)):
                return "date"
            return "text"

        columns = []
        if results:
            col_types: dict = {}
            for row in results:
                for key, value in row.items():
                    if key not in col_types or col_types[key] is None:
                        t = _py_to_col_type(value)
                        if t is not None:
                            col_types[key] = t
            for key in results[0].keys():
                col_types.setdefault(key, "text")
            columns = [{"name": k, "type": col_types[k]} for k in results[0].keys()]

        response = {
            "success": True,
            "data": results if (not limit or limit <= 0 or page) else results[:limit],
            "columns": columns,
            "total": total_count if total_count is not None else len(results),
            "origin": datasource.get('origin'),
            "executed_query": final_query
        }
        if page:
            response["page"] = page
            response["page_size"] = page_size
            response["total_pages"] = (total_count + page_size - 1) // page_size if total_count else 1

        return response
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"[ERROR] preview_unified_datasource: {e}")
        return {
            "success": False,
            "error": str(e),
            "data": [],
            "columns": []
        }


# =============================================================================
# DRILLDOWN GENERIQUE (pagination serveur, totaux, export)
# =============================================================================

def _normalize_ds_row(row: dict) -> dict:
    out = {}
    for k, v in row.items():
        if isinstance(v, Decimal):
            out[k] = float(v)
        elif isinstance(v, (datetime, date)):
            out[k] = v.isoformat()
        else:
            out[k] = v
    return out


def _resolve_dwh_for_ds(dwh_code, origin):
    effective = dwh_code
    if not effective and origin in ("template", "override"):
        try:
            dwh_list = execute_query(
                "SELECT TOP 1 code FROM APP_DWH WHERE actif = 1 ORDER BY id",
                use_cache=True,
            )
            if dwh_list:
                effective = dwh_list[0]["code"]
        except Exception:
            pass
    return effective


def _dedup_columns(sql):
    """Wrap a SELECT to deduplicate column names (SQL Server rejects duplicate cols in subqueries)."""
    match = re.match(r'(?is)(SELECT\s+(?:TOP\s+\d+\s+)?)(.*?)(FROM\s+.+)', sql)
    if not match:
        return sql
    prefix, cols_part, from_part = match.group(1), match.group(2), match.group(3)
    segments = []
    depth = 0
    current = []
    for ch in cols_part:
        if ch == '(':
            depth += 1
        elif ch == ')':
            depth -= 1
        if ch == ',' and depth == 0:
            segments.append(''.join(current).strip())
            current = []
        else:
            current.append(ch)
    if current:
        segments.append(''.join(current).strip())
    seen = {}
    deduped = []
    for seg in segments:
        m = re.search(r'(?:\[([^\]]+)\]|(\w+))\s*$', seg)
        alias = (m.group(1) or m.group(2)).strip() if m else seg.strip()
        if alias in seen:
            continue
        seen[alias] = True
        deduped.append(seg)
    return f"{prefix}{', '.join(deduped)} {from_part}"


def _extract_column_names(sql):
    """Extract column alias names from a SELECT query."""
    match = re.match(r'(?is)SELECT\s+(?:TOP\s+\d+\s+)?(.*?)FROM\s+', sql)
    if not match:
        return []
    cols_part = match.group(1)
    segments = []
    depth = 0
    current = []
    for ch in cols_part:
        if ch == '(':
            depth += 1
        elif ch == ')':
            depth -= 1
        if ch == ',' and depth == 0:
            segments.append(''.join(current).strip())
            current = []
        else:
            current.append(ch)
    if current:
        segments.append(''.join(current).strip())
    names = []
    for seg in segments:
        m = re.search(r'(?:\[([^\]]+)\]|(\w+))\s*$', seg)
        if m:
            names.append(m.group(1) or m.group(2))
    return names


def _resolve_filter_field(filter_field, column_names):
    """Map a filter field to the best matching column in the drilldown datasource."""
    if not filter_field or not column_names:
        return filter_field
    if filter_field in column_names:
        return filter_field
    fl = filter_field.lower()
    for col in column_names:
        if col.lower() == fl:
            return col
    for col in column_names:
        cl = col.lower()
        if fl in cl or cl in fl:
            return col
    COMMON_MAPPINGS = {
        "client": ["Intitulé client", "Code client", "Nom client"],
        "fournisseur": ["Intitulé fournisseur", "Code fournisseur", "Nom fournisseur"],
        "article": ["Désignation Article", "Code Article", "Référence article"],
        "famille": ["Intitulé famille", "Code Famille"],
        "commercial": ["Nom représentant", "Code représentant"],
        "depot": ["Dépôt", "Code dépôt"],
        "mois": ["Date BL", "Date"],
        "periode": ["Date BL", "Date"],
        "annee": ["Date BL", "Date"],
        "trimestre": ["Date BL", "Date"],
        "semaine": ["Date BL", "Date"],
        "region": ["Région", "Ville"],
    }
    for key, candidates in COMMON_MAPPINGS.items():
        if key in fl:
            for c in candidates:
                if c in column_names:
                    return c
    return filter_field


def _strip_sql_comments(sql):
    return re.sub(r'--[^\n]*', '', sql)


_TEMPORAL_FIELDS = {"periode", "mois", "annee", "trimestre", "semaine", "year", "month", "quarter", "week"}

_MONTH_NAMES_FR = {
    "janvier": 1, "février": 2, "mars": 3, "avril": 4, "mai": 5, "juin": 6,
    "juillet": 7, "août": 8, "aout": 8, "september": 9, "septembre": 9,
    "october": 10, "octobre": 10, "november": 11, "novembre": 11,
    "december": 12, "décembre": 12,
    "january": 1, "february": 2, "march": 3, "april": 4, "may": 5, "june": 6,
    "july": 7, "august": 8, "september": 9, "october": 10, "november": 11, "december": 12,
}


def _parse_temporal_filter(filter_field, filter_value, resolved_field):
    """Convert temporal filter values (e.g. 'August 2025', '2025-08', 'T3 2025') to SQL date range."""
    fl = filter_field.lower()
    val = str(filter_value).strip()
    sf = resolved_field.replace("]", "]]")

    if fl in ("annee", "year"):
        m = re.match(r'^(\d{4})$', val)
        if m:
            return f"YEAR([{sf}]) = ?", [int(m.group(1))]

    if fl in ("trimestre", "quarter"):
        m = re.match(r'^[TQ]?(\d)\s*[-/]?\s*(\d{4})$', val, re.IGNORECASE)
        if m:
            return f"DATEPART(QUARTER, [{sf}]) = ? AND YEAR([{sf}]) = ?", [int(m.group(1)), int(m.group(2))]
        m = re.match(r'^(\d{4})\s*[-/]?\s*[TQ](\d)$', val, re.IGNORECASE)
        if m:
            return f"YEAR([{sf}]) = ? AND DATEPART(QUARTER, [{sf}]) = ?", [int(m.group(1)), int(m.group(2))]

    # "August 2025", "Août 2025", "août 2025"
    m = re.match(r'^([A-Za-zÀ-ÿ]+)\s+(\d{4})$', val)
    if m:
        month_name = m.group(1).lower()
        year = int(m.group(2))
        month_num = _MONTH_NAMES_FR.get(month_name)
        if month_num:
            return f"MONTH([{sf}]) = ? AND YEAR([{sf}]) = ?", [month_num, year]

    # "2025-08" or "08/2025" or "08-2025"
    m = re.match(r'^(\d{4})[-/](\d{1,2})$', val)
    if m:
        return f"YEAR([{sf}]) = ? AND MONTH([{sf}]) = ?", [int(m.group(1)), int(m.group(2))]
    m = re.match(r'^(\d{1,2})[-/](\d{4})$', val)
    if m:
        return f"MONTH([{sf}]) = ? AND YEAR([{sf}]) = ?", [int(m.group(1)), int(m.group(2))]

    return None, None


def _build_drilldown_query(ds_query, context, filter_field, filter_value):
    from ..services.parameter_resolver import inject_params

    base_query = inject_params(ds_query, context or {})
    base_query = re.sub(r'\s+ORDER\s+BY\s+[\s\S]+$', '', base_query, flags=re.IGNORECASE)
    base_query = _strip_sql_comments(base_query)
    base_query = _dedup_columns(base_query)

    where_parts = []
    filter_params = []

    if filter_field and filter_value is not None:
        col_names = _extract_column_names(base_query)
        resolved_field = _resolve_filter_field(filter_field, col_names)
        field_exists = resolved_field in col_names or any(c.lower() == resolved_field.lower() for c in col_names)

        is_temporal = filter_field.lower() in _TEMPORAL_FIELDS
        if is_temporal and field_exists:
            temporal_clause, temporal_params = _parse_temporal_filter(filter_field, filter_value, resolved_field)
            if temporal_clause:
                where_parts.append(temporal_clause)
                filter_params.extend(temporal_params)
            else:
                safe_field = resolved_field.replace("]", "]]")
                where_parts.append(f"[{safe_field}] = ?")
                filter_params.append(filter_value)
        elif is_temporal and not field_exists:
            logger.warning(f"Drilldown: temporal field '{filter_field}' not resolved, skipping filter")
        elif field_exists:
            safe_field = resolved_field.replace("]", "]]")
            if str(filter_value) in ("None", ""):
                where_parts.append(f"[{safe_field}] IS NULL")
            else:
                where_parts.append(f"CAST([{safe_field}] AS NVARCHAR(MAX)) = ?")
                filter_params.append(str(filter_value))
        else:
            logger.warning(f"Drilldown: field '{filter_field}' not found in columns {col_names}, skipping filter")

    where_clause = " AND ".join(where_parts) if where_parts else "1=1"
    return base_query, where_clause, filter_params


@router.post("/unified/{identifier}/drilldown")
async def drilldown_datasource(
    identifier: str,
    request: DataSourceDrilldownRequest,
    x_dwh_code: str = Header(None, alias="X-DWH-Code"),
):
    """Drilldown generique sur une DataSource avec pagination serveur et totaux."""
    from ..database_unified import current_dwh_code as _ctx_dwh

    effective_header_dwh = x_dwh_code or _ctx_dwh.get()

    try:
        ds_response = await get_unified_datasource(identifier, effective_header_dwh)
        datasource = ds_response["data"]
        query = datasource.get("query_template", "")
        if not query:
            raise HTTPException(status_code=400, detail="Aucune requete definie")

        base_query, where_clause, filter_params = _build_drilldown_query(
            query, request.context, request.filterField, request.filterValue
        )

        page_size = request.pageSize or 50
        page = max(1, request.page or 1)
        offset = (page - 1) * page_size

        sort_field = request.sortField
        sort_direction = "DESC" if request.sortDirection and request.sortDirection.lower() == "desc" else "ASC"
        order_clause = f"ORDER BY [{sort_field}] {sort_direction}" if sort_field else "ORDER BY (SELECT NULL)"

        count_query = f"SELECT COUNT(*) AS __cnt__ FROM ({base_query}) AS __src__ WHERE {where_clause}"
        drill_query = f"""
            SELECT * FROM ({base_query}) AS __src__
            WHERE {where_clause}
            {order_clause}
            OFFSET {offset} ROWS FETCH NEXT {page_size} ROWS ONLY
        """

        effective_dwh = _resolve_dwh_for_ds(effective_header_dwh, datasource.get("origin"))
        params_tuple = tuple(filter_params) if filter_params else None

        if effective_dwh:
            count_rows = DWHConnectionManager.execute_dwh_query(effective_dwh, count_query, params_tuple, use_cache=False)
            page_data = DWHConnectionManager.execute_dwh_query(effective_dwh, drill_query, params_tuple, use_cache=False)
        else:
            count_rows = execute_query(count_query, params_tuple, use_cache=False)
            page_data = execute_query(drill_query, params_tuple, use_cache=False)

        total = count_rows[0]["__cnt__"] if count_rows else 0
        total_pages = max(1, (total + page_size - 1) // page_size)
        page_data = [_normalize_ds_row(r) for r in page_data]

        columns = []
        numeric_fields = []
        if page_data:
            for col_name in page_data[0].keys():
                col_info = {"field": col_name, "header": col_name}
                sample = page_data[0].get(col_name)
                if isinstance(sample, (int, float)):
                    col_info["format"] = "number"
                    numeric_fields.append(col_name)
                elif isinstance(sample, str) and len(sample) >= 10:
                    try:
                        datetime.fromisoformat(sample)
                        col_info["format"] = "date"
                    except (ValueError, TypeError):
                        pass
                columns.append(col_info)

        totals = {}
        if numeric_fields:
            sum_exprs = ", ".join(f"SUM(CAST([{f}] AS FLOAT)) AS [{f}]" for f in numeric_fields)
            totals_query = f"SELECT {sum_exprs} FROM ({base_query}) AS __src__ WHERE {where_clause}"
            if effective_dwh:
                tot_rows = DWHConnectionManager.execute_dwh_query(effective_dwh, totals_query, params_tuple, use_cache=False)
            else:
                tot_rows = execute_query(totals_query, params_tuple, use_cache=False)
            if tot_rows:
                row = _normalize_ds_row(tot_rows[0])
                totals = {k: v for k, v in row.items() if v is not None}

        return {
            "success": True,
            "data": page_data,
            "total": total,
            "page": page,
            "pageSize": page_size,
            "totalPages": total_pages,
            "columns": columns,
            "totals": totals,
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erreur drilldown datasource {identifier}: {e}")
        return {"success": False, "error": str(e)}


@router.post("/unified/{identifier}/drilldown/export")
async def drilldown_datasource_export(
    identifier: str,
    request: DataSourceDrilldownRequest,
    x_dwh_code: str = Header(None, alias="X-DWH-Code"),
):
    """Export Excel complet du drilldown datasource (toutes les lignes)."""
    from ..database_unified import current_dwh_code as _ctx_dwh

    effective_header_dwh = x_dwh_code or _ctx_dwh.get()

    try:
        import pandas as pd

        ds_response = await get_unified_datasource(identifier, effective_header_dwh)
        datasource = ds_response["data"]
        query = datasource.get("query_template", "")
        if not query:
            raise HTTPException(status_code=400, detail="Aucune requete definie")

        base_query, where_clause, filter_params = _build_drilldown_query(
            query, request.context, request.filterField, request.filterValue
        )

        sort_field = request.sortField
        sort_direction = "DESC" if request.sortDirection and request.sortDirection.lower() == "desc" else "ASC"
        order_clause = f"ORDER BY [{sort_field}] {sort_direction}" if sort_field else "ORDER BY (SELECT NULL)"

        export_query = f"""
            SELECT * FROM ({base_query}) AS __src__
            WHERE {where_clause}
            {order_clause}
        """

        effective_dwh = _resolve_dwh_for_ds(effective_header_dwh, datasource.get("origin"))
        params_tuple = tuple(filter_params) if filter_params else None

        if effective_dwh:
            all_data = DWHConnectionManager.execute_dwh_query(effective_dwh, export_query, params_tuple, use_cache=False)
        else:
            all_data = execute_query(export_query, params_tuple, use_cache=False)

        all_data = [_normalize_ds_row(r) for r in all_data]

        df = pd.DataFrame(all_data) if all_data else pd.DataFrame()

        if not df.empty:
            numeric_cols = df.select_dtypes(include="number").columns.tolist()
            totals_row = {
                col: df[col].sum() if col in numeric_cols else ("TOTAL" if col == df.columns[0] else "")
                for col in df.columns
            }
            df = pd.concat([df, pd.DataFrame([totals_row])], ignore_index=True)

        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Detail", index=False)
            ws = writer.sheets["Detail"]
            from openpyxl.styles import Font, PatternFill, Alignment

            header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
            total_fill = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            if ws.max_row > 1:
                for cell in ws[ws.max_row]:
                    cell.font = Font(bold=True)
                    cell.fill = total_fill
            for col in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

        buf.seek(0)
        safe_name = re.sub(r"[^a-zA-Z0-9_]", "_", identifier)
        filename = f"drilldown_{safe_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erreur export drilldown datasource {identifier}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# Mapping ASCII alias → label d'affichage avec accents corrects.
# Les alias SQL restent ASCII (evite la corruption pyodbc/SQL Server),
# mais l'UI affiche le francais complet via ce mapping.
_FIELD_LABELS = {
    # ── Quantités ──
    "Quantite":                     "Quantité",
    "Qte":                          "Qté",
    "Qte Totale":                   "Qté Totale",
    "Qte Vendue":                   "Qté Vendue",
    "Qte en Stock":                 "Qté en Stock",
    "Qte Entrees":                  "Qté Entrées",
    "Qte Sorties":                  "Qté Sorties",
    "Qte Achetee":                  "Qté Achetée",
    "Qte BC":                       "Qté BC",
    "Qte BL":                       "Qté BL",
    "Qte Commandee":                "Qté Commandée",
    "Qte Livree":                   "Qté Livrée",
    "Qte PL":                       "Qté PL",
    "Total Quantite":               "Total Quantité",
    # ── Périodes / dépôts ──
    "Periode":                      "Période",
    "Depot":                        "Dépôt",
    "Code Depot":                   "Code Dépôt",
    "Annee":                        "Année",
    "Trimestre":                    "Trimestre",
    "Semaine":                      "Semaine",
    "CA Annee N":                   "CA Année N",
    "CA Annee N-1":                 "CA Année N-1",
    "Annee N":                      "Année N",
    "Annee N-1":                    "Année N-1",
    # ── Échéances / recouvrement ──
    "Echeance":                     "Échéance",
    "Echeances":                    "Échéances",
    "Total Echeances":              "Total Échéances",
    "Nb Echeances":                 "Nb Échéances",
    "Nb Echeances +120j":           "Nb Échéances +120j",
    "Nb Echeances Retard":          "Nb Échéances Retard",
    "Derniere Echeance":            "Dernière Échéance",
    "Jours Avant Echeance":         "Jours Avant Échéance",
    "A Echoir":                     "À Échoir",
    "Echu":                         "Échu",
    "Total Echu":                   "Total Échu",
    "Nb Echues":                    "Nb Échues",
    "Non Echu":                     "Non Échu",
    # ── Règlements ──
    "Total Regle":                  "Total Réglé",
    "Montant Regle":                "Montant Réglé",
    "Regle 12 Mois":                "Réglé 12 Mois",
    "Reglements Mois":              "Règlements Mois",
    "Deja Regle":                   "Déjà Réglé",
    "Encaisse":                     "Encaissé",
    "Reste a Payer":                "Reste à Payer",
    "Reste A Livrer":               "Reste À Livrer",
    "Reste a Encaisser":            "Reste à Encaisser",
    # ── Créances / débiteurs ──
    "Creances":                     "Créances",
    "Creances Douteuses 120j":      "Créances Douteuses +120j",
    "Nb Clients Debiteurs":         "Nb Clients Débiteurs",
    "Solde Debiteur":               "Solde Débiteur",
    "Solde Crediteur":              "Solde Créditeur",
    "Solde Non Lettre":             "Solde Non Lettré",
    "Nb Lettrees":                  "Nb Lettrées",
    "Nb Non Lettrees":              "Nb Non Lettrées",
    # ── Mouvements stock ──
    "Valeur Entrees":               "Valeur Entrées",
    "Valeur Sorties":               "Valeur Sorties",
    # ── Libellés / désignations ──
    "Libelle":                      "Libellé",
    "Intitule":                     "Intitulé",
    "Intitule Compte":              "Intitulé Compte",
    "Designation":                  "Désignation",
    "Reference":                    "Référence",
    "Reference Client":             "Référence Client",
    # ── Catégories ──
    "Categorie":                    "Catégorie",
    "Categorie Tarifaire":          "Catégorie Tarifaire",
    # ── Résultats / délais ──
    "Resultat":                     "Résultat",
    "Resultat Net":                 "Résultat Net",
    "Resultat Exploitation":        "Résultat d'Exploitation",
    "Delai Moyen":                  "Délai Moyen",
    "Delai Moyen Jours":            "Délai Moyen Jours",
    "Delai BC vers BL (j)":         "Délai BC vers BL (j)",
    "Delai BL vers Facture (j)":    "Délai BL vers Facture (j)",
    "Delai Total BC vers Facture (j)": "Délai Total BC vers Facture (j)",
    "Max Delai BC-BL":              "Max Délai BC-BL",
    "Min Delai BC-BL":              "Min Délai BC-BL",
    # ── Divers ──
    "Premiere Vente":               "Première Vente",
    "Qualite":                      "Qualité",
    "Qualite Fournisseur":          "Qualité Fournisseur",
    "Tranche Age":                  "Tranche d'Âge",
    "Unite":                        "Unité",
    "Mvt Debit":                    "Mvt Débit",
    "Mvt Credit":                   "Mvt Crédit",
    "Nb Clients Retard":            "Nb Clients Retard",
}


@router.get("/unified/{identifier}/fields")
async def get_unified_datasource_fields(
    identifier: str,
    x_dwh_code: str = Header(None, alias="X-DWH-Code")
):
    """
    Recupere les champs (colonnes) d'une DataSource via cursor.description (fonctionne
    meme si la table est vide — SELECT TOP 0 retourne les meta-colonnes sans donnees).
    Résultat mis en cache 10 minutes.
    """
    from ..services.parameter_resolver import inject_params
    import re

    try:
        ds_response = await get_unified_datasource(identifier, x_dwh_code)
        datasource = ds_response['data']

        query = datasource.get('query_template', '')
        if not query:
            return {"success": True, "fields": []}

        # Clé de cache : identifier + dwh_code + début du template
        cache_key = hashlib.md5(
            f"{identifier}|{x_dwh_code or ''}|{query[:200]}".encode()
        ).hexdigest()
        cached = _UNIFIED_FIELDS_CACHE.get(cache_key)
        if cached and (time.time() - cached["ts"]) < _UNIFIED_FIELDS_TTL:
            return {"success": True, "fields": cached["fields"], "cached": True}

        # Injecter les parametres avec valeurs par defaut
        from ..services.parameter_resolver import get_default_context
        default_ctx = get_default_context()
        query = inject_params(query, default_ctx)

        # SELECT TOP 0 — retourne 0 lignes mais cursor.description contient les colonnes
        query = re.sub(r'\s+ORDER\s+BY\s+[\s\S]+$', '', query, flags=re.IGNORECASE)
        query = f"SELECT TOP 0 * FROM ({query}) AS __fields__"

        # Type codes pyodbc → type lisible
        import pyodbc
        from decimal import Decimal
        from datetime import datetime, date

        NUMBER_TYPES = {int, float, Decimal}

        def _run_and_extract(cursor):
            cursor.execute(query)
            if cursor.description is None:
                return []
            fields = []
            for col in cursor.description:
                col_name = col[0]
                # Heuristique sur le nom pour deviner le type quand impossible via valeurs
                name_lower = col_name.lower()
                if any(k in name_lower for k in ('date', 'echeance', 'livraison', 'facture', 'reglement')):
                    field_type = "date"
                elif any(k in name_lower for k in ('montant', 'ht', 'ttc', 'qte', 'quantit', 'prix', 'puht', 'puttc', 'taux', 'total', 'solde', 'credit', 'debit', 'poids', 'nb')):
                    field_type = "number"
                else:
                    field_type = "text"
                label = _FIELD_LABELS.get(col_name, col_name)
                fields.append({"name": col_name, "label": label, "type": field_type})
            return fields

        if x_dwh_code:
            with dwh_cursor(x_dwh_code) as cursor:
                fields = _run_and_extract(cursor)
        else:
            with get_db_cursor() as cursor:
                fields = _run_and_extract(cursor)

        # Injecter des champs derives Annee/Mois pour chaque colonne date/periode
        DATE_DERIVE_KEYWORDS = ('date', 'mois', 'periode', 'livraison', 'echeance', 'facture', 'reglement')
        derived = []
        for f in fields:
            name_lower = f['name'].lower()
            if f['type'] == 'date' or any(k in name_lower for k in DATE_DERIVE_KEYWORDS):
                src = f['name']
                src_label = f.get('label', src)
                derived.append({'name': src + '_Annee', 'label': f'Année ({src_label})', 'type': 'text', 'derived': True, 'source': src})
                derived.append({'name': src + '_Mois', 'label': f'Mois ({src_label})', 'type': 'text', 'derived': True, 'source': src})
        fields.extend(derived)

        _UNIFIED_FIELDS_CACHE[cache_key] = {"fields": fields, "ts": time.time()}
        return {"success": True, "fields": fields}
    except Exception as e:
        logger.error(f"[ERROR] get_unified_datasource_fields: {e}")
        return {"success": False, "error": str(e), "fields": []}


@router.get("/dwh-filter-options")
async def get_dwh_filter_options(
    field: str = Query(..., description="Champ à lister: societe, commercial, gamme, zone"),
    x_dwh_code: Optional[str] = Header(None, alias="X-DWH-Code")
):
    """
    Retourne les valeurs distinctes d'un champ depuis les données DWH.
    Utilisé pour alimenter les filtres globaux (GlobalFilterBar).
    """
    from ..database_unified import current_dwh_code as _ctx_dwh

    effective_dwh = x_dwh_code or _ctx_dwh.get()
    if not effective_dwh:
        return {"success": False, "error": "DWH code manquant", "data": []}

    # Mapping champ → colonne dans Lignes_des_ventes
    field_map = {
        "societe":    "[societe]",
        "commercial": "[Représentant]",
        "gamme":      "[Catalogue 1]",
        "zone":       "[Souche]",
    }
    column = field_map.get(field)
    if not column:
        raise HTTPException(status_code=400, detail=f"Champ non supporté: {field}. Valeurs: {list(field_map.keys())}")

    query = f"SELECT DISTINCT {column} AS value, {column} AS label FROM Lignes_des_ventes WHERE {column} IS NOT NULL ORDER BY {column}"

    try:
        results = DWHConnectionManager.execute_dwh_query(effective_dwh, query, use_cache=False)
        return {"success": True, "data": results}
    except Exception as e:
        logger.error(f"[ERROR] get_dwh_filter_options ({field}, {effective_dwh}): {e}")
        return {"success": False, "error": str(e), "data": []}


@router.post("/execute/test")
async def test_query(
    body: dict,
    x_user_role: str = Header(None, alias="X-User-Role")
):
    """Teste une requete SQL avec des parametres"""
    try:
        query = body.get("query", "")
        params = body.get("parameters", {})
        limit = body.get("limit", 10)

        if not query:
            raise HTTPException(status_code=400, detail="Requete vide")

        # Securite: verification basique
        query_lower = query.lower().strip()
        if any(kw in query_lower for kw in ['drop ', 'delete ', 'truncate ', 'insert ', 'update ', 'alter ', 'create ']):
            raise HTTPException(status_code=403, detail="Operations de modification non autorisees en mode test")

        # Remplacer les parametres @param par des valeurs de test (format dates sécurisé)
        from ..services.parameter_resolver import inject_params
        test_query = inject_params(query, params)

        # Ajouter TOP pour limiter les resultats
        if "SELECT" in test_query.upper() and "TOP" not in test_query.upper():
            test_query = test_query.replace("SELECT", f"SELECT TOP {limit}", 1)

        # Executer la requete
        results = execute_query(test_query, use_cache=False)

        # Extraire les colonnes
        columns = list(results[0].keys()) if results else []

        return {
            "success": True,
            "data": results[:limit],
            "rowCount": len(results),
            "columns": columns,
            "query_executed": test_query[:500] + "..." if len(test_query) > 500 else test_query
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"[ERROR] test_query: {e}")
        return {
            "success": False,
            "error": str(e),
            "data": [],
            "rowCount": 0,
            "columns": []
        }
