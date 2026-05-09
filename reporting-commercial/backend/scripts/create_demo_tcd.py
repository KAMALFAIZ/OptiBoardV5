"""
Script pour creer un classeur demo TCD et l'inserer dans la base.
Usage: python scripts/create_demo_tcd.py
"""
import sys, os, json, io, requests

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

API = "http://127.0.0.1:8085"

def create_demo_excel():
    wb = openpyxl.Workbook()
    thin = Side(style="thin", color="D0D0D0")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    alt_fill = PatternFill(start_color="EBF1F8", end_color="EBF1F8", fill_type="solid")
    num_fmt = "#,##0"

    data = [
        ["Sanitaire",   125000, 132000, 145000, 128000, 155000, 162000],
        ["Carrelage",     98000, 105000, 112000,  95000, 118000, 125000],
        ["Plomberie",     67000,  72000,  78000,  65000,  82000,  88000],
        ["Electricite",   45000,  48000,  52000,  43000,  55000,  58000],
        ["Outillage",     34000,  36000,  39000,  32000,  41000,  44000],
        ["Peinture",      28000,  30000,  33000,  27000,  35000,  37000],
    ]

    # ── Sheet 1: TCD Ventes par Famille / Mois ──
    ws1 = wb.active
    ws1.title = "TCD Ventes par Famille"

    ws1.merge_cells("A1:H1")
    ws1["A1"] = "TABLEAU CROISE DYNAMIQUE - Chiffre d'Affaires par Famille"
    ws1["A1"].font = Font(bold=True, size=13, color="1F4E79")
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 30

    ws1.merge_cells("A2:H2")
    ws1["A2"] = "Periode: Janvier - Juin 2026"
    ws1["A2"].font = Font(italic=True, size=10, color="666666")
    ws1["A2"].alignment = Alignment(horizontal="center")

    headers = ["Famille", "Jan", "Fev", "Mar", "Avr", "Mai", "Jun", "TOTAL"]
    for c, h in enumerate(headers, 1):
        cell = ws1.cell(row=4, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for r, row_data in enumerate(data):
        row_num = r + 5
        fill = alt_fill if r % 2 == 0 else None
        cell = ws1.cell(row=row_num, column=1, value=row_data[0])
        cell.font = Font(bold=True, size=10)
        cell.border = border
        if fill:
            cell.fill = fill
        for c, val in enumerate(row_data[1:], 2):
            cell = ws1.cell(row=row_num, column=c, value=val)
            cell.number_format = num_fmt
            cell.alignment = Alignment(horizontal="right")
            cell.border = border
            if fill:
                cell.fill = fill
        cell = ws1.cell(row=row_num, column=8, value=sum(row_data[1:]))
        cell.number_format = num_fmt
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal="right")
        cell.border = border
        if fill:
            cell.fill = fill

    total_row = len(data) + 5
    total_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    total_font = Font(bold=True, color="FFFFFF", size=10)

    cell = ws1.cell(row=total_row, column=1, value="TOTAL")
    cell.font = total_font
    cell.fill = total_fill
    cell.border = border
    grand = 0
    for c in range(2, 8):
        col_sum = sum(data[r][c - 1] for r in range(len(data)))
        grand += col_sum
        cell = ws1.cell(row=total_row, column=c, value=col_sum)
        cell.font = total_font
        cell.fill = total_fill
        cell.number_format = num_fmt
        cell.alignment = Alignment(horizontal="right")
        cell.border = border
    cell = ws1.cell(row=total_row, column=8, value=grand)
    cell.font = total_font
    cell.fill = total_fill
    cell.number_format = num_fmt
    cell.alignment = Alignment(horizontal="right")
    cell.border = border

    kpi_row = total_row + 2
    ws1.merge_cells(f"A{kpi_row}:B{kpi_row}")
    ws1[f"A{kpi_row}"] = "Evolution Moy/Mois:"
    ws1[f"A{kpi_row}"].font = Font(bold=True, size=10, color="1F4E79")
    ws1[f"C{kpi_row}"] = "+8.2%"
    ws1[f"C{kpi_row}"].font = Font(bold=True, size=12, color="27AE60")
    ws1[f"C{kpi_row}"].alignment = Alignment(horizontal="center")

    ws1.merge_cells(f"D{kpi_row}:E{kpi_row}")
    ws1[f"D{kpi_row}"] = "CA Cumule:"
    ws1[f"D{kpi_row}"].font = Font(bold=True, size=10, color="1F4E79")
    ca_total = sum(sum(row[1:]) for row in data)
    ws1[f"F{kpi_row}"] = ca_total
    ws1[f"F{kpi_row}"].font = Font(bold=True, size=12, color="1F4E79")
    ws1[f"F{kpi_row}"].number_format = "#,##0"

    ws1.column_dimensions["A"].width = 16
    for col in "BCDEFGH":
        ws1.column_dimensions[col].width = 14

    # ── Sheet 2: Top 10 Clients ──
    ws2 = wb.create_sheet("Top 10 Clients")

    ws2.merge_cells("A1:E1")
    ws2["A1"] = "TOP 10 CLIENTS - Semestre 1 2026"
    ws2["A1"].font = Font(bold=True, size=13, color="1F4E79")
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 30

    headers2 = ["Rang", "Client", "CA HT", "Nb Factures", "% du CA Total"]
    for c, h in enumerate(headers2, 1):
        cell = ws2.cell(row=3, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    clients = [
        [1, "MARJANE HOLDING",       485000, 45, "12.8%"],
        [2, "BIM STORES",            372000, 38, "9.8%"],
        [3, "ACIMA",                 298000, 32, "7.9%"],
        [4, "LABEL VIE",             265000, 28, "7.0%"],
        [5, "CARREFOUR MARKET",      234000, 25, "6.2%"],
        [6, "ASWAK ASSALAM",         198000, 22, "5.2%"],
        [7, "METRO CASH & CARRY",    176000, 20, "4.6%"],
        [8, "ATACADAO",              154000, 18, "4.1%"],
        [9, "KITEA",                 132000, 15, "3.5%"],
        [10, "MR BRICOLAGE",         118000, 12, "3.1%"],
    ]

    gold_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    silver_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
    bronze_fill = PatternFill(start_color="F0DCC4", end_color="F0DCC4", fill_type="solid")
    medal_fills = [gold_fill, silver_fill, bronze_fill]

    for r, row_data in enumerate(clients):
        row_num = r + 4
        fill = medal_fills[r] if r < 3 else (alt_fill if r % 2 == 0 else None)
        for c, val in enumerate(row_data, 1):
            cell = ws2.cell(row=row_num, column=c, value=val)
            cell.border = border
            if fill:
                cell.fill = fill
            if c == 1:
                cell.alignment = Alignment(horizontal="center")
                cell.font = Font(bold=True, size=10)
            elif c == 3:
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")
            elif c == 4:
                cell.alignment = Alignment(horizontal="center")
            elif c == 5:
                cell.alignment = Alignment(horizontal="center")
                cell.font = Font(bold=True, color="1F4E79")

    ws2.column_dimensions["A"].width = 8
    ws2.column_dimensions["B"].width = 24
    ws2.column_dimensions["C"].width = 14
    ws2.column_dimensions["D"].width = 14
    ws2.column_dimensions["E"].width = 14

    # ── Sheet 3: Evolution CA ──
    ws3 = wb.create_sheet("Evolution CA")

    ws3.merge_cells("A1:I1")
    ws3["A1"] = "EVOLUTION CA MENSUEL PAR FAMILLE"
    ws3["A1"].font = Font(bold=True, size=13, color="1F4E79")
    ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 30

    months = ["Jan", "Fev", "Mar", "Avr", "Mai", "Jun"]
    totals_by_month = [sum(d[i + 1] for d in data) for i in range(6)]
    bar_colors = ["4472C4", "5B9BD5", "70AD47", "FFC000", "ED7D31", "A5A5A5"]

    for c, (month, total) in enumerate(zip(months, totals_by_month)):
        col = c + 2
        cell = ws3.cell(row=3, column=col, value=month)
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal="center")
        cell = ws3.cell(row=4, column=col, value=total)
        cell.number_format = "#,##0"
        cell.alignment = Alignment(horizontal="center")
        cell.font = Font(bold=True, size=9, color=bar_colors[c])

    ws3.cell(row=6, column=1, value="Details par famille:").font = Font(bold=True, size=11, color="1F4E79")

    for c, h in enumerate(["Famille"] + months + ["Total", "Part %"], 1):
        cell = ws3.cell(row=7, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    grand_total = sum(sum(d[1:]) for d in data)
    for r, row_data in enumerate(data):
        row_num = r + 8
        fill = alt_fill if r % 2 == 0 else None
        cell = ws3.cell(row=row_num, column=1, value=row_data[0])
        cell.font = Font(bold=True)
        cell.border = border
        if fill:
            cell.fill = fill
        row_total = sum(row_data[1:])
        for c, val in enumerate(row_data[1:], 2):
            cell = ws3.cell(row=row_num, column=c, value=val)
            cell.number_format = "#,##0"
            cell.alignment = Alignment(horizontal="right")
            cell.border = border
            if fill:
                cell.fill = fill
        cell = ws3.cell(row=row_num, column=8, value=row_total)
        cell.number_format = "#,##0"
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="right")
        cell.border = border
        if fill:
            cell.fill = fill
        pct = f"{row_total / grand_total * 100:.1f}%"
        cell = ws3.cell(row=row_num, column=9, value=pct)
        cell.font = Font(bold=True, color="1F4E79")
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
        if fill:
            cell.fill = fill

    ws3.column_dimensions["A"].width = 16
    for col in "BCDEFGHI":
        ws3.column_dimensions[col].width = 13

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def main():
    print("1) Creation du fichier Excel demo TCD...")
    excel_buf = create_demo_excel()

    print("2) Import via API...")
    resp = requests.post(
        f"{API}/api/spreadsheet/import-excel",
        files={"file": ("Demo_TCD_Ventes.xlsx", excel_buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    if resp.status_code != 200:
        print(f"   ERREUR import: {resp.status_code} {resp.text[:300]}")
        return
    import_data = resp.json()
    if not import_data.get("success"):
        print(f"   ERREUR: {import_data}")
        return
    sheets = import_data["sheets"]
    print(f"   OK — {len(sheets)} feuilles importees")

    sheets_config = []
    for s in sheets:
        sheets_config.append({
            "name": s["name"],
            "data_source_code": None,
            "data_source_id": None,
            "column_mapping": [],
            "options": {},
            "imported_celldata": s["celldata"],
            "imported_config": s.get("config", {}),
            "imported_row_count": s.get("row_count", 0),
            "imported_column_count": s.get("column_count", 0),
        })

    print("3) Creation du classeur en base...")
    payload = {
        "nom": "Demo TCD — Analyse Ventes S1 2026",
        "description": "Tableau Croise Dynamique: CA par Famille/Mois, Top Clients, Evolution",
        "sheets": sheets_config,
        "features": {},
        "application": "commercial",
        "is_public": True,
    }
    resp = requests.post(f"{API}/api/spreadsheet/sheets", json=payload)
    if resp.status_code != 200:
        print(f"   ERREUR creation: {resp.status_code} {resp.text[:300]}")
        return
    result = resp.json()
    ss_id = result.get("id")
    print(f"   OK — Classeur cree avec id={ss_id}")

    print(f"\n   URL Viewer: /spreadsheet/{ss_id}")

    # 4) Add menu entry
    print("4) Ajout menu utilisateur...")
    import pyodbc
    try:
        conn = pyodbc.connect("DRIVER={ODBC Driver 17 for SQL Server};SERVER=localhost;DATABASE=OptiBoard_SaaS;Trusted_Connection=yes")
        cursor = conn.cursor()

        # Check APP_Menus columns
        cursor.execute("SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME='APP_Menus'")
        cols = [r[0] for r in cursor.fetchall()]
        print(f"   APP_Menus columns: {cols}")

        # Check if menu entry already exists
        cursor.execute("SELECT id FROM APP_Menus WHERE code = 'demo-tcd-ventes'")
        existing = cursor.fetchone()
        if existing:
            print(f"   Menu deja existant (id={existing[0]}), mise a jour...")
            cursor.execute(
                "UPDATE APP_Menus SET nom=?, url=?, actif=1 WHERE code='demo-tcd-ventes'",
                (f"Demo TCD Ventes", f"/spreadsheet/{ss_id}")
            )
        else:
            # Insert menu
            if "is_custom" in cols:
                cursor.execute(
                    "INSERT INTO APP_Menus (nom, code, icon, url, ordre, type, actif, is_custom) VALUES (?, ?, ?, ?, ?, ?, 1, 0)",
                    ("Demo TCD Ventes", "demo-tcd-ventes", "FileSpreadsheet", f"/spreadsheet/{ss_id}", 90, "page")
                )
            else:
                cursor.execute(
                    "INSERT INTO APP_Menus (nom, code, icon, url, ordre, type, actif) VALUES (?, ?, ?, ?, ?, ?, 1)",
                    ("Demo TCD Ventes", "demo-tcd-ventes", "FileSpreadsheet", f"/spreadsheet/{ss_id}", 90, "page")
                )
            print("   Menu cree dans APP_Menus (centrale)")

        conn.commit()
        conn.close()
    except Exception as e:
        print(f"   ERREUR menu: {e}")
        print(f"   -> Ajoutez manuellement le menu via /admin/menus avec URL: /spreadsheet/{ss_id}")

    print("\nTermine!")


if __name__ == "__main__":
    main()
