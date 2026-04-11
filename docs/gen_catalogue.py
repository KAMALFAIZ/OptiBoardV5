from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

wb = Workbook()

# ── COLOURS ──────────────────────────────────────────────────────────────────
HDR_BG   = "003366"
HDR_FG   = "FFFFFF"
TITLE_BG = "001F4D"
ROW_ODD  = "DDEEFF"
ROW_EVEN = "FFFFFF"

MODULE_COLORS = {
    "Ventes":             "C6EFCE",   # green
    "Recouvrement":       "FCE4D6",   # orange
    "Comptabilité":       "BDD7EE",   # blue
    "Stocks":             "E2EFDA",   # brown-green
    "Achats":             "E2D0F0",   # purple
    "Dettes Fournisseurs":"FFD7D7",   # red
    "Dashboard":          "D0F0F0",   # teal
    "Fiches":             "FFE8F5",   # pink
    "Analytique":         "FFFACD",   # yellow
    "Builder":            "EBEBEB",   # gray
}

TYPE_FONT = {
    "Dashboard KPI": {"bold": True},
    "Fiche":         {"italic": True},
}

thin = Side(style="thin", color="B0C4DE")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def hdr_cell(ws, row, col, value, bg=HDR_BG, fg=HDR_FG, sz=10, bold=True):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", bold=bold, color=fg, size=sz)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border
    return c


def data_cell(ws, row, col, value, bg=None, bold=False, italic=False, align="left"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", size=9, bold=bold, italic=italic)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    c.border = border
    return c


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — Catalogue Rapports
# ═══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Catalogue Rapports"

# Title row
ws1.merge_cells("A1:L1")
title = ws1["A1"]
title.value = "OPTIBOARD — CATALOGUE COMPLET DES RAPPORTS"
title.font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
title.fill = PatternFill("solid", fgColor=TITLE_BG)
title.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 28

# Headers row 2
HEADERS = [
    "N°", "Module", "Catégorie", "Nom du Rapport", "Description",
    "Type", "Endpoint API", "Source de données", "Filtres disponibles",
    "Export", "Drill-down", "Statut"
]
for ci, h in enumerate(HEADERS, 1):
    hdr_cell(ws1, 2, ci, h)
ws1.row_dimensions[2].height = 30

# ── DATA ──────────────────────────────────────────────────────────────────────
reports = [
    # (module, catégorie, nom, description, type, endpoint, source, filtres, export, drill)
    ("Ventes","commercial","CA par Gamme","Chiffre d'affaires ventilé par famille de produits avec comparatif N/N-1","Pivot","GET /api/ventes/par-gamme","Lignes_des_ventes + Entete_des_ventes","Période / Société / Commercial / Zone","Oui","Oui"),
    ("Ventes","commercial","CA par Commercial","Performance des représentants : CA, marge, taux de marge","Pivot","GET /api/ventes/par-commercial","Lignes_des_ventes + Entete_des_ventes","Période / Société / Gamme","Oui","Oui"),
    ("Ventes","commercial","Top 100 Clients","Classement des meilleurs clients par CA décroissant","GridView","GET /api/ventes/top-clients","Lignes_des_ventes + Entete_des_ventes + Clients","Période / Société / Gamme / Commercial","Oui","Oui"),
    ("Ventes","commercial","Top 100 Produits","Classement des articles les plus vendus par CA et quantité","GridView","GET /api/ventes/top-produits","Lignes_des_ventes + Articles","Période / Société / Gamme / Canal","Oui","Oui"),
    ("Ventes","commercial","CA par Période","Évolution mensuelle et annuelle du chiffre d'affaires","Dashboard KPI","GET /api/ventes","Lignes_des_ventes + Entete_des_ventes","Période / Société / Gamme / Canal / Zone / Commercial","Oui","Non"),
    ("Ventes","commercial","Liste des Ventes","Liste agrégée des ventes avec tous les axes d'analyse disponibles","GridView","GET /api/liste-ventes","Lignes_des_ventes + Entete_des_ventes","Société / Gamme / Catalogue / Commercial / Canal / Zone / Région / Ville / Groupe client","Oui","Non"),
    ("Ventes","commercial","Factures de Ventes","Liste de toutes les factures clients avec statut règlement","GridView","GET /api/gridview (DS_VTE_FACTURES)","Lignes_des_ventes + Entete_des_ventes","Période / Client / Commercial / Société","Oui","Non"),
    ("Ventes","commercial","Bons de Livraison","Liste des bons de livraison émis","GridView","GET /api/gridview (DS_VTE_BL)","Lignes_des_ventes + Entete_des_ventes","Période / Client / Dépôt","Oui","Non"),
    ("Ventes","commercial","Bons de Commande Ventes","Commandes clients en cours ou traitées","GridView","GET /api/gridview (DS_VTE_BC)","Lignes_des_ventes + Entete_des_ventes","Période / Client / Commercial","Oui","Non"),
    ("Ventes","commercial","Devis","Propositions commerciales émises aux clients","GridView","GET /api/gridview (DS_VTE_DEVIS)","Lignes_des_ventes + Entete_des_ventes","Période / Client / Commercial","Oui","Non"),
    ("Ventes","commercial","Avoirs Ventes","Avoirs clients émis pour retours ou corrections","GridView","GET /api/gridview (DS_VTE_AVOIRS)","Lignes_des_ventes + Entete_des_ventes","Période / Client","Oui","Non"),
    ("Ventes","commercial","Bons de Retour Ventes","Retours marchandises clients","GridView","GET /api/gridview (DS_VTE_RETOURS)","Lignes_des_ventes + Entete_des_ventes","Période / Client","Oui","Non"),
    ("Ventes","commercial","Commandes Préparées","Commandes en cours de préparation en entrepôt","GridView","GET /api/gridview (DS_VTE_PREPAREES)","Lignes_des_ventes + Entete_des_ventes","Période / Dépôt","Oui","Non"),
    ("Ventes","commercial","Détail par Gamme","Drill-down : lignes de vente pour une gamme de produit donnée","GridView","GET /api/ventes/detail/gamme/{gamme}","Lignes_des_ventes + Entete_des_ventes","Gamme / Période","Oui","Non"),
    ("Ventes","commercial","Détail par Client","Historique des achats d'un client spécifique","GridView","GET /api/ventes/detail/client/{code_client}","Lignes_des_ventes + Entete_des_ventes + Clients","Client / Période","Oui","Non"),
    ("Ventes","commercial","Détail par Produit","Historique des ventes d'un article spécifique","GridView","GET /api/ventes/detail/produit/{code_article}","Lignes_des_ventes + Articles","Article / Période","Oui","Non"),
    ("Ventes","commercial","Détail par Commercial","Lignes de ventes réalisées par un représentant","GridView","GET /api/ventes/detail/commercial/{commercial}","Lignes_des_ventes + Entete_des_ventes","Commercial / Période","Oui","Non"),
    ("Ventes","commercial","Détail par Mois","Lignes de ventes pour un mois et une année spécifiques","GridView","GET /api/ventes/detail/mois/{annee}/{mois}","Lignes_des_ventes + Entete_des_ventes","Année / Mois","Oui","Non"),
    ("Recouvrement","recouvrement","Tableau de bord Recouvrement","KPIs globaux : encours total, DSO, taux impayés, alertes","Dashboard KPI","GET /api/recouvrement","Echeances_Ventes + Clients","Société / Commercial / Période","Non","Non"),
    ("Recouvrement","recouvrement","DSO (Days Sales Outstanding)","Calcul du délai moyen de règlement clients","Dashboard KPI","GET /api/recouvrement/dso","Echeances_Ventes + Lignes_des_ventes","Période / Société","Non","Non"),
    ("Recouvrement","recouvrement","Balance Âgée Clients","Créances classées par tranches d'ancienneté (0-30, 31-60, 61-90, 91-120, +120 jours)","Pivot","GET /api/recouvrement/balance-agee","Echeances_Ventes + Clients","Société / Commercial / Client","Oui","Oui"),
    ("Recouvrement","recouvrement","Fiche Client Recouvrement","Détail complet d'un client : encours, échéances, historique règlements","Fiche","GET /api/recouvrement/client/{client_id}","Echeances_Ventes + Règlements_Clients + Lignes_des_ventes","Client","Oui","Non"),
    ("Recouvrement","recouvrement","Recouvrement par Commercial","Encours et impayés agrégés par représentant","Pivot","GET /api/recouvrement/commercial/{commercial_id}","Echeances_Ventes + Clients","Commercial / Période","Oui","Oui"),
    ("Recouvrement","recouvrement","Échéances par Tranche","Détail des créances pour une tranche d'ancienneté spécifique","GridView","GET /api/recouvrement/tranche/{tranche}","Echeances_Ventes + Clients","Tranche / Société","Oui","Non"),
    ("Recouvrement","recouvrement","Échéances Clients","Liste complète des échéances clients non réglées","GridView","GET /api/recouvrement/echeances","Echeances_Ventes + Imputation_Factures_Ventes","Période / Client / Commercial / Mode règlement","Oui","Non"),
    ("Recouvrement","recouvrement","Échéances par Client (agrégé)","Montants dus agrégés par client","GridView","GET /api/recouvrement/echeances/par-client","Echeances_Ventes + Clients","Période / Société","Oui","Oui"),
    ("Recouvrement","recouvrement","Échéances par Commercial (agrégé)","Montants dus agrégés par représentant","GridView","GET /api/recouvrement/echeances/par-commercial","Echeances_Ventes + Clients","Période","Oui","Oui"),
    ("Recouvrement","recouvrement","Échéances par Mode de Règlement","Répartition des créances par moyen de paiement","GridView","GET /api/recouvrement/echeances/par-mode-reglement","Echeances_Ventes","Période / Société","Oui","Non"),
    ("Recouvrement","recouvrement","Échéances à Échoir","Créances dont l'échéance est à venir","GridView","GET /api/recouvrement/echeances/a-echoir","Echeances_Ventes + Clients","Société","Oui","Non"),
    ("Recouvrement","recouvrement","Règlements Clients","Historique de tous les paiements reçus","GridView","GET /api/recouvrement/reglements","Règlements_Clients","Période / Client","Oui","Non"),
    ("Recouvrement","recouvrement","Règlements par Client","Paiements agrégés par client","GridView","GET /api/recouvrement/reglements/par-client","Règlements_Clients + Clients","Période","Oui","Oui"),
    ("Recouvrement","recouvrement","Règlements par Mode","Paiements agrégés par moyen de paiement","GridView","GET /api/recouvrement/reglements/par-mode","Règlements_Clients","Période","Oui","Non"),
    ("Recouvrement","recouvrement","Factures Non Réglées","Liste des factures dont le solde reste impayé","GridView","GET /api/recouvrement/factures-non-reglees","Imputation_Factures_Ventes + Echeances_Ventes","Période / Client / Commercial","Oui","Non"),
    ("Recouvrement","recouvrement","Historique Client","Historique complet des transactions et règlements d'un client","Fiche","GET /api/recouvrement/historique-client/{code_client}","Echeances_Ventes + Règlements_Clients + Lignes_des_ventes","Client / Période","Oui","Non"),
    ("Recouvrement","recouvrement","Évolution Recouvrement","Tendance du recouvrement dans le temps (encours, règlements, DSO)","Dashboard KPI","GET /api/recouvrement/evolution","Echeances_Ventes + Règlements_Clients","Période / Société","Non","Non"),
    ("Comptabilité","comptabilite","KPIs Comptables","Indicateurs clés : résultat net, trésorerie, encours clients, dettes fournisseurs","Dashboard KPI","GET /api/comptabilite/kpis","Ecritures_Comptables","Période / Société / Exercice","Non","Non"),
    ("Comptabilité","comptabilite","Balance Générale","Balance de tous les comptes avec soldes débiteurs/créditeurs","GridView","GET /api/comptabilite/balance-generale","Ecritures_Comptables","Période / Exercice / Compte","Oui","Oui"),
    ("Comptabilité","comptabilite","Journal des Écritures","Grand livre de toutes les écritures comptables","GridView","GET /api/comptabilite/journal-ecritures","Ecritures_Comptables","Période / Journal / Compte","Oui","Non"),
    ("Comptabilité","comptabilite","Balance Tiers","Balance clients et fournisseurs avec soldes","GridView","GET /api/comptabilite/balance-tiers","Ecritures_Comptables","Période / Type tiers (client/fournisseur)","Oui","Oui"),
    ("Comptabilité","comptabilite","Trésorerie","Écritures de trésorerie et soldes bancaires","GridView","GET /api/comptabilite/tresorerie","Ecritures_Comptables","Période / Banque / Journal","Oui","Non"),
    ("Comptabilité","comptabilite","Détail des Charges","Charges ventilées par nature de compte","GridView","GET /api/comptabilite/charges","Ecritures_Comptables","Période / Catégorie compte","Oui","Oui"),
    ("Comptabilité","comptabilite","Détail des Produits Comptables","Produits ventilés par nature de compte","GridView","GET /api/comptabilite/produits","Ecritures_Comptables","Période / Catégorie compte","Oui","Oui"),
    ("Comptabilité","comptabilite","Échéances Clients Comptables","Échéances ventes non réglées issues de la comptabilité","GridView","GET /api/comptabilite/echeances-clients","Ecritures_Comptables","Période / Client","Oui","Non"),
    ("Comptabilité","comptabilite","Échéances Fournisseurs Comptables","Échéances achats à payer issues de la comptabilité","GridView","GET /api/comptabilite/echeances-fournisseurs","Ecritures_Comptables","Période / Fournisseur","Oui","Non"),
    ("Comptabilité","comptabilite","Lettrage","Rapprochement des écritures lettrées et non lettrées","GridView","GET /api/comptabilite/lettrage","Ecritures_Comptables","Période / Compte / Journal","Oui","Non"),
    ("Comptabilité","comptabilite","Analyses Comptables","Évolution mensuelle des comptes de résultat et de bilan","Pivot","GET /api/comptabilite/analyses","Ecritures_Comptables","Exercice / Type analyse","Oui","Non"),
    ("Stocks","stocks","État Global des Stocks","Vue d'ensemble des stocks : valeur, quantités, alertes rupture","Dashboard KPI","GET /api/stocks","Etat_Stock","Dépôt / Gamme / Catégorie","Non","Non"),
    ("Stocks","stocks","Articles Dormants","Articles sans mouvement depuis plus de 180 jours","GridView","GET /api/stocks/dormant","Etat_Stock + Mouvement_stock","Dépôt / Gamme / Seuil jours","Oui","Non"),
    ("Stocks","stocks","Rotation par Gamme","Taux de rotation du stock par famille de produits","Pivot","GET /api/stocks/rotation","Mouvement_stock + Etat_Stock","Période / Gamme / Dépôt","Oui","Oui"),
    ("Stocks","stocks","Détail Stock par Article","Stock détaillé d'un article : quantité, valeur, historique mouvements","Fiche","GET /api/stocks/article/{code_article}","Etat_Stock + Mouvement_stock","Article / Dépôt","Oui","Non"),
    ("Stocks","stocks","Stock par Gamme","Stocks agrégés par famille de produits","GridView","GET /api/stocks/par-gamme","Etat_Stock","Gamme / Dépôt","Oui","Oui"),
    ("Stocks","stocks","Mouvements de Stock","Historique de tous les mouvements (entrées/sorties/transferts)","GridView","GET /api/gridview (DS_STK_MOUVEMENTS)","Mouvement_stock","Période / Type mouvement / Dépôt / Article","Oui","Non"),
    ("Stocks","stocks","Entrées de Stock","Mouvements d'entrée uniquement (réceptions, achats, retours)","GridView","GET /api/gridview (DS_STK_ENTREES)","Mouvement_stock","Période / Dépôt / Fournisseur","Oui","Non"),
    ("Stocks","stocks","Sorties de Stock","Mouvements de sortie uniquement (ventes, consommations, pertes)","GridView","GET /api/gridview (DS_STK_SORTIES)","Mouvement_stock","Période / Dépôt / Client","Oui","Non"),
    ("Stocks","stocks","État Actuel du Stock","Snapshot en temps réel des quantités disponibles par article/dépôt","GridView","GET /api/gridview (DS_STK_ETAT_ACTUEL)","Etat_Stock","Dépôt / Gamme / Famille","Oui","Non"),
    ("Stocks","stocks","Stock par Dépôt","Répartition des stocks entre les différents dépôts/entrepôts","Pivot","GET /api/gridview (DS_STK_PAR_DEPOT)","Etat_Stock","Dépôt / Gamme","Oui","Oui"),
    ("Stocks","stocks","Articles en Rupture","Articles dont le stock est inférieur au seuil de réapprovisionnement","GridView","GET /api/gridview (DS_STK_RUPTURE)","Etat_Stock","Dépôt / Gamme / Seuil","Oui","Non"),
    ("Achats","achats","Factures Fournisseurs","Liste de toutes les factures fournisseurs reçues","GridView","GET /api/gridview (DS_ACH_FACTURES)","Lignes_des_achats + Entête_des_achats","Période / Fournisseur / Famille","Oui","Non"),
    ("Achats","achats","Bons de Livraison Achats","Réceptions fournisseurs","GridView","GET /api/gridview (DS_ACH_BL)","Lignes_des_achats + Entête_des_achats","Période / Fournisseur","Oui","Non"),
    ("Achats","achats","Bons de Commande Achats","Commandes passées aux fournisseurs","GridView","GET /api/gridview (DS_ACH_BC)","Lignes_des_achats + Entête_des_achats","Période / Fournisseur","Oui","Non"),
    ("Achats","achats","Demandes d'Achat","Demandes internes d'approvisionnement","GridView","GET /api/gridview (DS_ACH_DA)","Lignes_des_achats + Entête_des_achats","Période / Demandeur","Oui","Non"),
    ("Achats","achats","Avoirs Achats","Avoirs reçus des fournisseurs pour retours ou corrections","GridView","GET /api/gridview (DS_ACH_AVOIRS)","Lignes_des_achats + Entête_des_achats","Période / Fournisseur","Oui","Non"),
    ("Achats","achats","Retours Achats","Retours de marchandises aux fournisseurs","GridView","GET /api/gridview (DS_ACH_RETOURS)","Lignes_des_achats + Entête_des_achats","Période / Fournisseur","Oui","Non"),
    ("Achats","achats","CA Achats par Fournisseur","Montants achetés agrégés par fournisseur","Pivot","GET /api/gridview (DS_ACH_PAR_FOURNISSEUR)","Lignes_des_achats + Entête_des_achats + Fournisseurs","Période / Famille","Oui","Oui"),
    ("Achats","achats","Achats par Article","Achats agrégés par article avec quantités et montants","Pivot","GET /api/gridview (DS_ACH_PAR_ARTICLE)","Lignes_des_achats + Articles","Période / Fournisseur","Oui","Oui"),
    ("Achats","achats","Achats par Famille","Achats agrégés par famille de produits","Pivot","GET /api/gridview (DS_ACH_PAR_FAMILLE)","Lignes_des_achats + Articles","Période / Fournisseur","Oui","Oui"),
    ("Dettes Fournisseurs","dettes_fournisseurs","Tableau de bord Dettes","KPIs globaux : dettes totales, échéances à venir, alertes","Dashboard KPI","GET /api/dettes-fournisseurs","Echeances_Achats + Fournisseurs","Société / Période","Non","Non"),
    ("Dettes Fournisseurs","dettes_fournisseurs","Balance Âgée Fournisseurs","Dettes classées par tranches d'ancienneté","Pivot","GET /api/dettes-fournisseurs/balance-agee","Echeances_Achats + Fournisseurs","Société / Fournisseur","Oui","Oui"),
    ("Dettes Fournisseurs","dettes_fournisseurs","Échéances Fournisseurs","Liste des échéances fournisseurs à payer","GridView","GET /api/dettes-fournisseurs/echeances","Echeances_Achats + Imputation_Factures_Achats","Période / Fournisseur / Mode règlement","Oui","Non"),
    ("Dettes Fournisseurs","dettes_fournisseurs","Dettes Impayées","Factures fournisseurs dont le règlement est en retard","GridView","GET /api/dettes-fournisseurs/impayees","Echeances_Achats + Imputation_Factures_Achats","Période / Fournisseur","Oui","Non"),
    ("Dettes Fournisseurs","dettes_fournisseurs","Dettes par Fournisseur","Encours dettes agrégé par fournisseur","GridView","GET /api/dettes-fournisseurs/par-fournisseur","Echeances_Achats + Fournisseurs","Période","Oui","Oui"),
    ("Dettes Fournisseurs","dettes_fournisseurs","Dettes par Mode de Règlement","Répartition des dettes par moyen de paiement","GridView","GET /api/dettes-fournisseurs/par-mode","Echeances_Achats","Période","Oui","Non"),
    ("Dashboard","dashboard","Dashboard Principal","Vue consolidée : CA, balance âgée, DSO, alertes, comparatif N/N-1","Dashboard KPI","GET /api/dashboard","Lignes_des_ventes + Echeances_Ventes + Ecritures_Comptables","Période / Société","Non","Oui"),
    ("Dashboard","dashboard","Analyse CA + Créances KPIs","Tableau de bord combiné : CA et créances clients en un seul écran","Dashboard KPI","GET /api/analyse-ca-creances/kpis","Lignes_des_ventes + Echeances_Ventes + Clients","Société / Commercial / Région","Non","Oui"),
    ("Dashboard","dashboard","Top Clients CA + Créances","Matrice : meilleurs clients par CA croisé avec leur encours","GridView","GET /api/analyse-ca-creances/top-clients-ca","Lignes_des_ventes + Echeances_Ventes + Clients","Période / Commercial","Oui","Oui"),
    ("Dashboard","dashboard","Top Clients Créances","Classement des clients avec les encours les plus élevés","GridView","GET /api/analyse-ca-creances/top-clients-creances","Echeances_Ventes + Clients","Période","Oui","Oui"),
    ("Dashboard","dashboard","CA par Mois (Dashboard)","Courbe d'évolution mensuelle du CA pour le dashboard","Dashboard KPI","GET /api/analyse-ca-creances/ca-par-mois","Lignes_des_ventes + Entete_des_ventes","Exercice / Société","Non","Non"),
    ("Dashboard","dashboard","CA par Commercial (Dashboard)","Histogramme CA par représentant pour le dashboard","Dashboard KPI","GET /api/analyse-ca-creances/ca-par-commercial","Lignes_des_ventes + Entete_des_ventes","Période / Société","Non","Oui"),
    ("Dashboard","dashboard","Balance Âgée par Tranche (Dashboard)","Jauge de répartition des créances par ancienneté","Dashboard KPI","GET /api/analyse-ca-creances/balance-agee-tranche","Echeances_Ventes + Clients","Société","Non","Oui"),
    ("Dashboard","dashboard","Balance Âgée Détail (Dashboard)","Tableau détaillé créances par client et tranche pour le dashboard","GridView","GET /api/analyse-ca-creances/balance-agee-detail","Echeances_Ventes + Clients","Société / Commercial","Oui","Non"),
    ("Fiches","client360","Fiche Client","Vue 360° client : infos, CA, encours, historique transactions","Fiche","GET /api/fiche-client/{code}","Clients + Lignes_des_ventes + Echeances_Ventes","Client","Oui","Oui"),
    ("Fiches","client360","Fiche Fournisseur","Vue 360° fournisseur : infos, achats, dettes, historique","Fiche","GET /api/fiche-fournisseur/{code}","Fournisseurs + Lignes_des_achats + Echeances_Achats","Fournisseur","Oui","Oui"),
    ("Analytique","commercial","Forecasting / Prévisions","Projections de ventes basées sur historique et tendances","Dashboard KPI","GET /api/forecast","Lignes_des_ventes + Entete_des_ventes","Période / Gamme / Commercial","Oui","Non"),
    ("Analytique","commercial","Détection d'Anomalies","Identification automatique des transactions anormales","Dashboard KPI","GET /api/anomalies","Lignes_des_ventes + Echeances_Ventes","Période / Type anomalie","Oui","Non"),
    ("Analytique","commercial","Alertes","Règles d'alerte configurables : seuils CA, créances, stock","Dashboard KPI","GET /api/alerts","Multiple","Module / Type alerte","Non","Non"),
    ("Analytique","commercial","Drillthrough","Drill-through générique depuis n'importe quelle cellule pivot vers le détail","GridView","GET /api/drillthrough","Dynamique (selon cellule source)","Dynamiques","Oui","Non"),
    ("Builder","builder","Pivot Builder V2","Constructeur de rapports pivot personnalisés avec glisser-déposer","Pivot","GET /api/v2/pivots","Toutes sources configurées","Configurables","Oui","Oui"),
    ("Builder","builder","GridView Builder","Constructeur de grilles de données avec colonnes et formatage personnalisables","GridView","GET /api/gridview","Toutes sources configurées","Configurables","Oui","Non"),
    ("Builder","builder","Planificateur de Rapports","Envoi automatique de rapports par email selon planning récurrent","Planificateur","GET /api/report-scheduler","N/A (orchestration)","Rapport / Destinataires / Fréquence","Non","Non"),
]

for i, r in enumerate(reports):
    row = i + 3
    module = r[0]
    typ = r[4]
    bg = MODULE_COLORS.get(module, "FFFFFF")
    tf = TYPE_FONT.get(typ, {})
    bold = tf.get("bold", False)
    italic = tf.get("italic", False)

    data_cell(ws1, row, 1, i + 1, bg=bg, align="center")
    data_cell(ws1, row, 2, module, bg=bg, bold=True)
    data_cell(ws1, row, 3, r[1], bg=bg)
    data_cell(ws1, row, 4, r[2], bg=bg, bold=True)
    data_cell(ws1, row, 5, r[3], bg=bg)
    data_cell(ws1, row, 6, r[4], bg=bg, bold=bold, italic=italic)
    data_cell(ws1, row, 7, r[5], bg=bg)
    data_cell(ws1, row, 8, r[6], bg=bg)
    data_cell(ws1, row, 9, r[7], bg=bg)
    export_bg = "C6EFCE" if r[8] == "Oui" else "FFD7D7"
    data_cell(ws1, row, 10, r[8], bg=export_bg, align="center")
    drill_bg = "C6EFCE" if r[9] == "Oui" else "FFD7D7"
    data_cell(ws1, row, 11, r[9], bg=drill_bg, align="center")
    data_cell(ws1, row, 12, "Actif", bg="C6EFCE", align="center")

    ws1.row_dimensions[row].height = 38

# Column widths
col_widths = [5, 18, 16, 28, 52, 14, 42, 42, 40, 8, 10, 8]
for ci, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(ci)].width = w

# Freeze panes
ws1.freeze_panes = "A3"

# Auto-filter on row 2
ws1.auto_filter.ref = f"A2:{get_column_letter(len(HEADERS))}2"


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — Résumé par Module
# ═══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Résumé par Module")

from collections import defaultdict

module_stats = defaultdict(lambda: {"GridView": 0, "Pivot": 0, "Dashboard KPI": 0, "Fiche": 0, "Autres": 0})
for r in reports:
    module = r[0]
    typ = r[4]
    if typ in ("GridView", "Pivot", "Dashboard KPI", "Fiche"):
        module_stats[module][typ] += 1
    else:
        module_stats[module]["Autres"] += 1

ws2.merge_cells("A1:H1")
t = ws2["A1"]
t.value = "OPTIBOARD — RÉSUMÉ PAR MODULE"
t.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
t.fill = PatternFill("solid", fgColor=TITLE_BG)
t.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 26

s2_headers = ["Module", "GridView", "Pivot", "Dashboard KPI", "Fiche", "Planificateur", "Total"]
for ci, h in enumerate(s2_headers, 1):
    hdr_cell(ws2, 2, ci, h)
ws2.row_dimensions[2].height = 24

MODULES_ORDER = ["Ventes", "Recouvrement", "Comptabilité", "Stocks", "Achats",
                 "Dettes Fournisseurs", "Dashboard", "Fiches", "Analytique", "Builder"]

for ri, mod in enumerate(MODULES_ORDER):
    row = ri + 3
    s = module_stats[mod]
    autres = s["Autres"]
    bg = MODULE_COLORS.get(mod, "FFFFFF")
    total = s["GridView"] + s["Pivot"] + s["Dashboard KPI"] + s["Fiche"] + autres
    vals = [mod, s["GridView"], s["Pivot"], s["Dashboard KPI"], s["Fiche"], autres, total]
    for ci, v in enumerate(vals, 1):
        c = ws2.cell(row=row, column=ci, value=v)
        c.font = Font(name="Arial", size=10, bold=(ci == 1 or ci == 7))
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center" if ci > 1 else "left", vertical="center")
        c.border = border

# Total row
total_row = len(MODULES_ORDER) + 3
ws2.cell(total_row, 1, "TOTAL").font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
ws2.cell(total_row, 1).fill = PatternFill("solid", fgColor=HDR_BG)
ws2.cell(total_row, 1).border = border
ws2.cell(total_row, 1).alignment = Alignment(horizontal="left", vertical="center")
for ci in range(2, 8):
    col_letter = get_column_letter(ci)
    formula = f"=SUM({col_letter}3:{col_letter}{total_row - 1})"
    c = ws2.cell(total_row, ci, formula)
    c.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=HDR_BG)
    c.border = border
    c.alignment = Alignment(horizontal="center", vertical="center")

for ci, w in enumerate([22, 12, 10, 16, 10, 14, 10], 1):
    ws2.column_dimensions[get_column_letter(ci)].width = w
ws2.freeze_panes = "A3"


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — Résumé par Type
# ═══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Résumé par Type")

ws3.merge_cells("A1:E1")
t3 = ws3["A1"]
t3.value = "OPTIBOARD — RÉSUMÉ PAR TYPE DE RAPPORT"
t3.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
t3.fill = PatternFill("solid", fgColor=TITLE_BG)
t3.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 26

s3_headers = ["Type de Rapport", "Nombre", "% du Total", "Modules concernés", "Description"]
for ci, h in enumerate(s3_headers, 1):
    hdr_cell(ws3, 2, ci, h)

type_stats = defaultdict(list)
for r in reports:
    type_stats[r[4]].append(r[0])

TYPE_COLORS = {
    "GridView":      "BDD7EE",
    "Pivot":         "C6EFCE",
    "Dashboard KPI": "FFE699",
    "Fiche":         "FFD7D7",
    "Planificateur": "E2D0F0",
}
TYPE_DESC = {
    "GridView":      "Tableau de données tabulaire avec tri, filtres, pagination et export",
    "Pivot":         "Analyse multidimensionnelle avec axes configurables (lignes/colonnes/valeurs)",
    "Dashboard KPI": "Indicateurs clés visuels : KPIs, graphiques, jauges, comparatifs",
    "Fiche":         "Vue 360° détaillée d'une entité (client, fournisseur, article)",
    "Planificateur": "Automatisation de l'envoi de rapports selon un planning récurrent",
}

total_reports = len(reports)
for ri, (typ, mods) in enumerate(sorted(type_stats.items(), key=lambda x: -len(x[1])), 1):
    row = ri + 2
    count = len(mods)
    unique_mods = ", ".join(sorted(set(mods)))
    bg = TYPE_COLORS.get(typ, "FFFFFF")
    vals = [typ, count, f"=B{row}/B{len(type_stats)+3}", unique_mods, TYPE_DESC.get(typ, "")]
    for ci, v in enumerate(vals, 1):
        c = ws3.cell(row=row, column=ci, value=v)
        c.font = Font(name="Arial", size=10, bold=(ci == 1))
        c.fill = PatternFill("solid", fgColor=bg)
        c.border = border
        c.alignment = Alignment(horizontal="center" if ci in (2, 3) else "left", vertical="center", wrap_text=True)
        if ci == 3:
            c.number_format = "0.0%"
    ws3.row_dimensions[row].height = 22

# Total
total_r3 = len(type_stats) + 3
ws3.cell(total_r3, 1, "TOTAL").font = Font(name="Arial", bold=True, color="FFFFFF")
ws3.cell(total_r3, 1).fill = PatternFill("solid", fgColor=HDR_BG)
ws3.cell(total_r3, 1).border = border
ws3.cell(total_r3, 1).alignment = Alignment(horizontal="left", vertical="center")
c_tot = ws3.cell(total_r3, 2, f"=SUM(B3:B{total_r3-1})")
c_tot.font = Font(name="Arial", bold=True, color="FFFFFF")
c_tot.fill = PatternFill("solid", fgColor=HDR_BG)
c_tot.border = border
c_tot.alignment = Alignment(horizontal="center", vertical="center")
c_pct = ws3.cell(total_r3, 3, 1)
c_pct.number_format = "0.0%"
c_pct.font = Font(name="Arial", bold=True, color="FFFFFF")
c_pct.fill = PatternFill("solid", fgColor=HDR_BG)
c_pct.border = border
c_pct.alignment = Alignment(horizontal="center", vertical="center")

for ci, w in enumerate([22, 10, 12, 50, 55], 1):
    ws3.column_dimensions[get_column_letter(ci)].width = w
ws3.freeze_panes = "A3"

# ── SAVE ──────────────────────────────────────────────────────────────────────
out = r"D:\kasoft-platform\OptiBoard\docs\OptiBoard_Rapports_Catalogue.xlsx"
wb.save(out)
print(f"Saved: {out}")
